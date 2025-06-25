# ===============================
# INICIO DEL PROGRAMA PRINCIPAL
# ===============================
# Este archivo es el punto de entrada de la aplicación de escritorio CNT-91.
# Al ejecutarlo, se carga la ventana principal con el diseño profesional definido en frontend_widgets.py

import tkinter as tk
from tkinter import ttk
from frontend_widgets import crear_layout_principal, get_info_cnt91_sections, get_info_cnt91_resources
import CNT_9X_pendulum as CNT
import os
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np

# Variable global para el objeto del frecuencímetro
cnt_device = None

# Valor por defecto del identificador GPIB
DEFAULT_GPIB = 'GPIB0::10::INSTR'

# Variable global para el canal seleccionado ('A' o 'B')
canal_seleccionado = 'A'

# Variable global para el intervalo de captura (en segundos)
intervalo_s = 0.2

# Variable global para el acoplamiento ('AC' o 'DC')
acoplamiento = 'AC'

# Variable global para la impedancia ('Max' o 'Min')
impedancia = 'Max'  # Nuevo parámetro: 'Max' para 1MΩ, 'Min' para 50Ω

# Variable global para la atenuación ('1' o '10')
atenuacion = '1'  # Nuevo parámetro: '1' para 1x, '10' para 10x

# Variable global para el nivel de trigger (None para automático, float para manual)
trigger_level = None  # Nuevo parámetro: None para automático, float para manual (-5 a 5)

# Variable global para la pendiente del trigger ('POS' o 'NEG')
trigger_slope = 'POS'  # Nuevo parámetro: 'POS' para pendiente positiva, 'NEG' para negativa

# Variable global para el filtro analógico ajustable (None para False, 'True' para True)
filtro_Analog_PASSAbaja = None  # Nuevo parámetro: None para False, 'True' para True

# Variable global para rastrear si se ha guardado la configuración
configuracion_guardada = False

# Variable global para rastrear si la medición está pausada
medicion_pausada = False

# Variable global para almacenar la ruta del archivo Excel
ruta_archivo_excel = None

# Función para mostrar el menú de selección de canal
def mostrar_menu_canal(widgets):
    global canal_seleccionado, intervalo_s, acoplamiento, impedancia, atenuacion, trigger_level, trigger_slope, filtro_Analog_PASSAbaja, configuracion_guardada, medicion_pausada, ruta_archivo_excel
    frame_contenido = widgets['frame_contenido']
    
    # Resetear el estado de configuración guardada
    configuracion_guardada = False
    
    # Resetear el estado de medición pausada
    medicion_pausada = False
    
    # Resetear la ruta del archivo Excel
    ruta_archivo_excel = None
    
    # Limpiar el frame de contenido
    for widget in frame_contenido.winfo_children():
        widget.destroy()
    
    # Frame principal con padding
    main_frame = tk.Frame(frame_contenido, bg='#f6f7fa')
    main_frame.pack(fill='both', expand=True, padx=10, pady=5)
    
    # Título principal - Frequency Datalogger
    titulo_frame = tk.Frame(main_frame, bg='#f6f7fa')
    titulo_frame.pack(fill='x', pady=(0, 10))
    
    titulo = tk.Label(titulo_frame, text='Frequency Datalogger', 
                     font=('Segoe UI', 14, 'bold'), 
                     fg='#25364a', bg='#f6f7fa')
    titulo.pack(anchor='w')
    
    # Línea separadora
    separador = tk.Frame(titulo_frame, height=1, bg='#2980f2')
    separador.pack(fill='x', pady=(3, 0))
    
    # Frame para las pestañas
    tabs_frame = tk.Frame(main_frame, bg='white', relief='solid', bd=1)
    tabs_frame.pack(fill='both', expand=True)
    
    # Frame superior para los botones de pestañas
    tabs_buttons_frame = tk.Frame(tabs_frame, bg='#f8f9fa', height=40)
    tabs_buttons_frame.pack(fill='x')
    tabs_buttons_frame.pack_propagate(False)
    
    # Frame para el contenido de las pestañas
    tabs_content_frame = tk.Frame(tabs_frame, bg='white')
    tabs_content_frame.pack(fill='both', expand=True)
    
    # Variables para controlar las pestañas
    current_tab = tk.StringVar(value='config')
    
    # Función para cambiar entre pestañas
    def switch_tab(tab_name):
        current_tab.set(tab_name)
        # Ocultar todos los frames de contenido
        for widget in tabs_content_frame.winfo_children():
            widget.pack_forget()
        
        # Mostrar el frame correspondiente
        if tab_name == 'config':
            config_content_frame.pack(fill='both', expand=True)
            btn_config_tab.config(bg='#2980f2', fg='white')
            btn_datalogger_tab.config(bg='#f8f9fa', fg='#25364a')
        else:
            datalogger_content_frame.pack(fill='both', expand=True)
            btn_config_tab.config(bg='#f8f9fa', fg='#25364a')
            btn_datalogger_tab.config(bg='#2980f2', fg='white')
    
    # Botones de pestañas
    btn_config_tab = tk.Button(tabs_buttons_frame, text='Configuración', 
                              command=lambda: switch_tab('config'),
                              font=('Segoe UI', 10, 'bold'),
                              bg='#2980f2', fg='white',
                              relief='flat', padx=20, pady=8,
                              cursor='hand2')
    btn_config_tab.pack(side='left', padx=(10, 2))
    
    btn_datalogger_tab = tk.Button(tabs_buttons_frame, text='Datalogger', 
                                  command=lambda: switch_tab('datalogger'),
                                  font=('Segoe UI', 10, 'bold'),
                                  bg='#f8f9fa', fg='#25364a',
                                  relief='flat', padx=20, pady=8,
                                  cursor='hand2')
    btn_datalogger_tab.pack(side='left', padx=(2, 10))
    
    # ===== PESTAÑA DE CONFIGURACIÓN =====
    config_content_frame = tk.Frame(tabs_content_frame, bg='white')
    
    # Título de configuración
    config_titulo = tk.Label(config_content_frame, text='Configuración:', 
                            font=('Segoe UI', 10, 'bold'), 
                            fg='#25364a', bg='white')
    config_titulo.pack(anchor='w', padx=10, pady=(8, 5))
    
    # Variable interna para el selector de canal
    canal_var = tk.StringVar(value='A')
    
    # Variable interna para el intervalo
    intervalo_var = tk.StringVar(value=str(intervalo_s))
    
    # Variable interna para el acoplamiento
    acoplamiento_var = tk.StringVar(value=acoplamiento)
    
    # Variable interna para la impedancia
    impedancia_var = tk.StringVar(value=impedancia)
    
    # Variable interna para la atenuación
    atenuacion_var = tk.StringVar(value=atenuacion)
    
    # Variable interna para el trigger
    trigger_mode_var = tk.StringVar(value='automatico' if trigger_level is None else 'manual')
    trigger_value_var = tk.StringVar(value=str(trigger_level) if trigger_level is not None else '0')
    
    # Variable interna para la pendiente del trigger
    trigger_slope_var = tk.StringVar(value=trigger_slope)
    
    # Variable interna para el filtro analógico
    filtro_analog_var = tk.StringVar(value='True' if filtro_Analog_PASSAbaja == 'True' else 'False')
    
    # Frame para el selector de canal
    selector_frame = tk.Frame(config_content_frame, bg='white')
    selector_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de canal
    label_canal = tk.Label(selector_frame, text='Canal de medición:', 
                          font=('Segoe UI', 8, 'bold'), 
                          fg='#6c757d', bg='white')
    label_canal.pack(anchor='w', pady=(0, 3))
    
    # Frame para los radio buttons de canal
    radio_frame = tk.Frame(selector_frame, bg='white')
    radio_frame.pack(anchor='w')
    
    # Radio buttons con mejor estilo
    radio_a = tk.Radiobutton(radio_frame, text='Canal A', 
                            variable=canal_var, value='A', 
                            font=('Segoe UI', 8), 
                            fg='#25364a', bg='white',
                            selectcolor='white',
                            activebackground='white',
                            activeforeground='#2980f2')
    radio_a.pack(side='left', padx=(0, 15))
    
    radio_b = tk.Radiobutton(radio_frame, text='Canal B', 
                            variable=canal_var, value='B', 
                            font=('Segoe UI', 8), 
                            fg='#25364a', bg='white',
                            selectcolor='white',
                            activebackground='white',
                            activeforeground='#2980f2')
    radio_b.pack(side='left')
    
    # Separador entre configuraciones
    separador_config = tk.Frame(config_content_frame, height=1, bg='#e0e0e0')
    separador_config.pack(fill='x', padx=10, pady=5)
    
    # Frame para el selector de intervalo
    intervalo_frame = tk.Frame(config_content_frame, bg='white')
    intervalo_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de intervalo
    label_intervalo = tk.Label(intervalo_frame, text='Intervalo de captura (segundos):', 
                              font=('Segoe UI', 8, 'bold'), 
                              fg='#6c757d', bg='white')
    label_intervalo.pack(anchor='w', pady=(0, 3))
    
    # Etiqueta informativa de límites y valor por defecto
    label_intervalo_info = tk.Label(intervalo_frame, text='(2E-8 a 1000 s, por defecto 0.2 s)', 
                                   font=('Segoe UI', 7), 
                                   fg='#6c757d', bg='white')
    label_intervalo_info.pack(anchor='w', pady=(0, 3))
    
    # Frame para el control de intervalo
    control_intervalo_frame = tk.Frame(intervalo_frame, bg='white')
    control_intervalo_frame.pack(anchor='w')
    
    # Función para validar entrada manual
    def validar_entrada_intervalo(*args):
        try:
            texto_actual = intervalo_var.get()
            if texto_actual == "" or texto_actual == "0":
                return  # Permitir campo vacío o solo "0"
            
            valor = float(texto_actual)
            
            # Solo validar si el valor está completo y es menor que el mínimo
            if valor < 2e-8 and valor != 0:
                intervalo_var.set("2.00e-08")
            elif valor > 1000:
                intervalo_var.set("1000.000")
        except ValueError:
            # Si no es un número válido, no hacer nada (permitir que el usuario siga escribiendo)
            pass
    
    # Vincular validación a cambios en la variable
    intervalo_var.trace('w', validar_entrada_intervalo)
    
    # Entry para el valor del intervalo
    entry_intervalo = tk.Entry(control_intervalo_frame, 
                              textvariable=intervalo_var,
                              font=('Segoe UI', 8),
                              width=12,
                              justify='left')
    entry_intervalo.pack(side='left')
    
    # Separador entre configuraciones
    separador_config2 = tk.Frame(config_content_frame, height=1, bg='#e0e0e0')
    separador_config2.pack(fill='x', padx=10, pady=5)
    
    # Frame para el selector de acoplamiento
    acoplamiento_frame = tk.Frame(config_content_frame, bg='white')
    acoplamiento_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de acoplamiento
    label_acoplamiento = tk.Label(acoplamiento_frame, text='Acoplamiento:', 
                                 font=('Segoe UI', 8, 'bold'), 
                                 fg='#6c757d', bg='white')
    label_acoplamiento.pack(anchor='w', pady=(0, 3))
    
    # Frame para los radio buttons de acoplamiento
    radio_acoplamiento_frame = tk.Frame(acoplamiento_frame, bg='white')
    radio_acoplamiento_frame.pack(anchor='w')
    
    # Radio buttons para acoplamiento
    radio_ac = tk.Radiobutton(radio_acoplamiento_frame, text='AC', 
                             variable=acoplamiento_var, value='AC', 
                             font=('Segoe UI', 8), 
                             fg='#25364a', bg='white',
                             selectcolor='white',
                             activebackground='white',
                             activeforeground='#2980f2')
    radio_ac.pack(side='left', padx=(0, 15))
    
    radio_dc = tk.Radiobutton(radio_acoplamiento_frame, text='DC', 
                             variable=acoplamiento_var, value='DC', 
                             font=('Segoe UI', 8), 
                             fg='#25364a', bg='white',
                             selectcolor='white',
                             activebackground='white',
                             activeforeground='#2980f2')
    radio_dc.pack(side='left')
    
    # Separador entre configuraciones
    separador_config3 = tk.Frame(config_content_frame, height=1, bg='#e0e0e0')
    separador_config3.pack(fill='x', padx=10, pady=5)
    
    # Frame para el selector de impedancia
    impedancia_frame = tk.Frame(config_content_frame, bg='white')
    impedancia_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de impedancia
    label_impedancia = tk.Label(impedancia_frame, text='Impedancia:', 
                                 font=('Segoe UI', 8, 'bold'), 
                                 fg='#6c757d', bg='white')
    label_impedancia.pack(anchor='w', pady=(0, 3))
    
    # Frame para los radio buttons de impedancia
    radio_impedancia_frame = tk.Frame(impedancia_frame, bg='white')
    radio_impedancia_frame.pack(anchor='w')
    
    # Radio buttons para impedancia
    radio_max = tk.Radiobutton(radio_impedancia_frame, text='1MΩ (Max)', 
                             variable=impedancia_var, value='Max', 
                             font=('Segoe UI', 8), 
                             fg='#25364a', bg='white',
                             selectcolor='white',
                             activebackground='white',
                             activeforeground='#2980f2')
    radio_max.pack(side='left', padx=(0, 15))
    
    radio_min = tk.Radiobutton(radio_impedancia_frame, text='50Ω (Min)', 
                             variable=impedancia_var, value='Min', 
                             font=('Segoe UI', 8), 
                             fg='#25364a', bg='white',
                             selectcolor='white',
                             activebackground='white',
                             activeforeground='#2980f2')
    radio_min.pack(side='left')
    
    # Separador entre configuraciones
    separador_config4 = tk.Frame(config_content_frame, height=1, bg='#e0e0e0')
    separador_config4.pack(fill='x', padx=10, pady=5)
    
    # Frame para el selector de atenuación
    atenuacion_frame = tk.Frame(config_content_frame, bg='white')
    atenuacion_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de atenuación
    label_atenuacion = tk.Label(atenuacion_frame, text='Atenuación:', 
                                 font=('Segoe UI', 8, 'bold'), 
                                 fg='#6c757d', bg='white')
    label_atenuacion.pack(anchor='w', pady=(0, 3))
    
    # Frame para los radio buttons de atenuación
    radio_atenuacion_frame = tk.Frame(atenuacion_frame, bg='white')
    radio_atenuacion_frame.pack(anchor='w')
    
    # Radio buttons para atenuación
    radio_1x = tk.Radiobutton(radio_atenuacion_frame, text='1x', 
                             variable=atenuacion_var, value='1', 
                             font=('Segoe UI', 8), 
                             fg='#25364a', bg='white',
                             selectcolor='white',
                             activebackground='white',
                             activeforeground='#2980f2')
    radio_1x.pack(side='left', padx=(0, 15))
    
    radio_10x = tk.Radiobutton(radio_atenuacion_frame, text='10x', 
                             variable=atenuacion_var, value='10', 
                             font=('Segoe UI', 8), 
                             fg='#25364a', bg='white',
                             selectcolor='white',
                             activebackground='white',
                             activeforeground='#2980f2')
    radio_10x.pack(side='left')
    
    # Separador entre configuraciones
    separador_config5 = tk.Frame(config_content_frame, height=1, bg='#e0e0e0')
    separador_config5.pack(fill='x', padx=10, pady=5)
    
    # Frame para el selector de trigger
    trigger_frame = tk.Frame(config_content_frame, bg='white')
    trigger_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de trigger
    label_trigger = tk.Label(trigger_frame, text='Nivel de Trigger:', 
                            font=('Segoe UI', 8, 'bold'), 
                            fg='#6c757d', bg='white')
    label_trigger.pack(anchor='w', pady=(0, 3))
    
    # Etiqueta explicativa de porcentajes automáticos
    def actualizar_explicacion_trigger(*args):
        canal_actual = canal_var.get()
        if canal_actual == 'A':
            porcentaje = '70%'
        else:
            porcentaje = '30%'
        explicacion_trigger.config(text=f"(Automático: {porcentaje} del rango)")
    
    explicacion_trigger = tk.Label(trigger_frame, text='(Automático: 70% del rango)', 
                                  font=('Segoe UI', 7), 
                                  fg='#6c757d', bg='white')
    explicacion_trigger.pack(anchor='w', pady=(0, 3))
    
    # Vincular actualización de explicación a cambios en el canal
    canal_var.trace('w', actualizar_explicacion_trigger)
    
    # Frame para los radio buttons de trigger
    radio_trigger_frame = tk.Frame(trigger_frame, bg='white')
    radio_trigger_frame.pack(anchor='w')
    
    # Radio buttons para trigger
    radio_auto = tk.Radiobutton(radio_trigger_frame, text='Automático', 
                               variable=trigger_mode_var, value='automatico', 
                               font=('Segoe UI', 8), 
                               fg='#25364a', bg='white',
                               selectcolor='white',
                               activebackground='white',
                               activeforeground='#2980f2')
    radio_auto.pack(side='left', padx=(0, 15))
    
    radio_manual = tk.Radiobutton(radio_trigger_frame, text='Manual', 
                                 variable=trigger_mode_var, value='manual', 
                                 font=('Segoe UI', 8), 
                                 fg='#25364a', bg='white',
                                 selectcolor='white',
                                 activebackground='white',
                                 activeforeground='#2980f2')
    radio_manual.pack(side='left')
    
    # Frame para el valor manual del trigger
    trigger_value_frame = tk.Frame(trigger_frame, bg='white')
    trigger_value_frame.pack(anchor='w', pady=(5, 0))
    
    # Etiqueta del valor manual
    label_trigger_value = tk.Label(trigger_value_frame, text='Valor (-5 a 5 V):', 
                                  font=('Segoe UI', 8), 
                                  fg='#6c757d', bg='white')
    label_trigger_value.pack(side='left', padx=(0, 5))
    
    # Entry para el valor manual del trigger
    entry_trigger = tk.Entry(trigger_value_frame, 
                            textvariable=trigger_value_var,
                            font=('Segoe UI', 8),
                            width=8,
                            justify='left')
    entry_trigger.pack(side='left')
    
    # Función para validar entrada del trigger
    def validar_entrada_trigger(*args):
        try:
            texto_actual = trigger_value_var.get()
            if texto_actual == "":
                return  # Permitir campo vacío
            
            valor = float(texto_actual)
            
            # Validar límites
            if valor < -5:
                trigger_value_var.set("-5.0")
            elif valor > 5:
                trigger_value_var.set("5.0")
        except ValueError:
            # Si no es un número válido, no hacer nada
            pass
    
    # Vincular validación a cambios en la variable
    trigger_value_var.trace('w', validar_entrada_trigger)
    
    # Separador entre configuraciones
    separador_config6 = tk.Frame(config_content_frame, height=1, bg='#e0e0e0')
    separador_config6.pack(fill='x', padx=10, pady=5)
    
    # Frame para el selector de pendiente del trigger
    trigger_slope_frame = tk.Frame(config_content_frame, bg='white')
    trigger_slope_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de pendiente del trigger
    label_trigger_slope = tk.Label(trigger_slope_frame, text='Pendiente del Trigger:', 
                                  font=('Segoe UI', 8, 'bold'), 
                                  fg='#6c757d', bg='white')
    label_trigger_slope.pack(anchor='w', pady=(0, 3))
    
    # Frame para los radio buttons de pendiente del trigger
    radio_trigger_slope_frame = tk.Frame(trigger_slope_frame, bg='white')
    radio_trigger_slope_frame.pack(anchor='w')
    
    # Radio buttons para pendiente del trigger
    radio_pos = tk.Radiobutton(radio_trigger_slope_frame, text='Positiva', 
                              variable=trigger_slope_var, value='POS', 
                              font=('Segoe UI', 8), 
                              fg='#25364a', bg='white',
                              selectcolor='white',
                              activebackground='white',
                              activeforeground='#2980f2')
    radio_pos.pack(side='left', padx=(0, 15))
    
    radio_neg = tk.Radiobutton(radio_trigger_slope_frame, text='Negativa', 
                              variable=trigger_slope_var, value='NEG', 
                              font=('Segoe UI', 8), 
                              fg='#25364a', bg='white',
                              selectcolor='white',
                              activebackground='white',
                              activeforeground='#2980f2')
    radio_neg.pack(side='left')
    
    # Separador entre configuraciones
    separador_config7 = tk.Frame(config_content_frame, height=1, bg='#e0e0e0')
    separador_config7.pack(fill='x', padx=10, pady=5)
    
    # Frame para el selector de filtro analógico
    filtro_analog_frame = tk.Frame(config_content_frame, bg='white')
    filtro_analog_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de filtro analógico
    label_filtro_analog = tk.Label(filtro_analog_frame, text='Filtro Analógico pasa bajas:', 
                                  font=('Segoe UI', 8, 'bold'), 
                                  fg='#6c757d', bg='white')
    label_filtro_analog.pack(anchor='w', pady=(0, 3))
    
    # Frame para los radio buttons de filtro analógico
    radio_filtro_analog_frame = tk.Frame(filtro_analog_frame, bg='white')
    radio_filtro_analog_frame.pack(anchor='w')
    
    # Radio buttons para filtro analógico
    radio_filtro_true = tk.Radiobutton(radio_filtro_analog_frame, text='True', 
                                      variable=filtro_analog_var, value='True', 
                                      font=('Segoe UI', 8), 
                                      fg='#25364a', bg='white',
                                      selectcolor='white',
                                      activebackground='white',
                                      activeforeground='#2980f2')
    radio_filtro_true.pack(side='left', padx=(0, 15))
    
    radio_filtro_false = tk.Radiobutton(radio_filtro_analog_frame, text='False', 
                                       variable=filtro_analog_var, value='False', 
                                       font=('Segoe UI', 8), 
                                       fg='#25364a', bg='white',
                                       selectcolor='white',
                                       activebackground='white',
                                       activeforeground='#2980f2')
    radio_filtro_false.pack(side='left')
    
    # Función para guardar selección
    def guardar_seleccion():
        global canal_seleccionado, intervalo_s, acoplamiento, impedancia, atenuacion, trigger_level, trigger_slope, filtro_Analog_PASSAbaja, configuracion_guardada, ruta_archivo_excel
        canal_seleccionado = canal_var.get()
        acoplamiento = acoplamiento_var.get()
        impedancia = impedancia_var.get()
        atenuacion = atenuacion_var.get()
        trigger_slope = trigger_slope_var.get()
        
        # Procesar filtro analógico
        if filtro_analog_var.get() == 'True':
            filtro_Analog_PASSAbaja = 'True'
        else:
            filtro_Analog_PASSAbaja = None
        
        # Procesar trigger
        if trigger_mode_var.get() == 'automatico':
            trigger_level = None
        else:
            try:
                trigger_level = float(trigger_value_var.get())
                # Validar límites finales
                if trigger_level < -5:
                    trigger_level = -5.0
                elif trigger_level > 5:
                    trigger_level = 5.0
            except ValueError:
                trigger_level = 0.0  # Valor por defecto si hay error
        
        try:
            intervalo_s = float(intervalo_var.get())
            # Validar límites finales
            if intervalo_s < 2e-8:
                intervalo_s = 2e-8
            elif intervalo_s > 1000:
                intervalo_s = 1000
        except ValueError:
            intervalo_s = 0.2  # Valor por defecto si hay error
        
        # Formatear el intervalo para mostrar
        if intervalo_s < 0.001:
            intervalo_formato = f"{intervalo_s:.2e}"
        elif intervalo_s < 1:
            intervalo_formato = f"{intervalo_s:.6f}"
        else:
            intervalo_formato = f"{intervalo_s:.3f}"
        
        # Formatear trigger para mostrar
        if trigger_level is None:
            trigger_formato = "Automático"
        else:
            trigger_formato = f"{trigger_level:.1f}V"
        
        # Formatear filtro analógico para mostrar
        filtro_formato = "True" if filtro_Analog_PASSAbaja == 'True' else "False"
        
        resultado_label.config(text=f"✓ Configuración guardada: Canal {canal_seleccionado}, Intervalo {intervalo_formato} s, Acoplamiento {acoplamiento}, Impedancia {impedancia}, Atenuación {atenuacion}x, Trigger {trigger_formato}, Pendiente {trigger_slope}, Filtro Analógico pasa bajas {filtro_formato}", fg='#27ae60')
        
        # Convertir filtro_Analog_PASSAbaja a booleano real
        filtro_analog_bool = True if filtro_Analog_PASSAbaja == 'True' else False
        
        # Configurar el dispositivo usando la función configurar_dispositivo
        try:
            # Acceder al objeto CNT_91 global (cnt_device)
            if 'cnt_device' in globals() and cnt_device is not None:
                # Convertir impedancia de 'Max'/'Min' a 'MAX'/'MIN' para la función
                impedancia_convertida = 'MAX' if impedancia == 'Max' else 'MIN'
                
                # Llamar a configurar_dispositivo con todos los parámetros
                file_path = cnt_device.configurar_dispositivo(
                    canal=canal_seleccionado,
                    intervalo_s=intervalo_s,
                    acoplamiento=acoplamiento,
                    impedancia=impedancia_convertida,
                    atenuacion=atenuacion,
                    trigger_level=trigger_level,
                    trigger_slope=trigger_slope,
                    filtro_Digital_PASSAbaja=None,
                    filtro_Analog_PASSAbaja=filtro_analog_bool,
                    file_path=None
                )
                
                # Actualizar la ruta del archivo Excel
                ruta_archivo_excel = file_path
                ruta_label.config(text=f'Guardando en: {file_path}', fg='#27ae60')
                
                # Marcar configuración como guardada y habilitar botón de datalogger
                configuracion_guardada = True
                btn_start_stop.config(state='normal', bg='#27ae60', fg='white', cursor='hand2')
                btn_fin_medicion.config(state='normal')
                status_label.config(text='Estado: Listo para iniciar', fg='#27ae60')
                
            else:
                resultado_label.config(text="⚠ Error: Dispositivo no conectado", fg='#e74c3c')
        except Exception as e:
            resultado_label.config(text=f"⚠ Error al configurar dispositivo: {str(e)}", fg='#e74c3c')
        
        # Actualizar la información de configuración en la pestaña de Datalogger
        actualizar_info_configuracion()
    
    # Frame para botón y resultado
    accion_frame = tk.Frame(config_content_frame, bg='white')
    accion_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Botón para guardar selección con estilo profesional
    btn_guardar = tk.Button(accion_frame, text='Guardar Configuración', 
                           command=guardar_seleccion, 
                           font=('Segoe UI', 8, 'bold'),
                           bg='#2980f2', fg='white',
                           relief='flat', padx=10, pady=3,
                           cursor='hand2')
    btn_guardar.pack(side='left', pady=(3, 0))
    
    # Label para mostrar resultado
    resultado_label = tk.Label(accion_frame, text='', 
                              font=('Segoe UI', 8), 
                              fg='#27ae60', bg='white')
    resultado_label.pack(side='left', padx=(8, 0), pady=(3, 0))
    
    # ===== PESTAÑA DE DATALOGGER =====
    datalogger_content_frame = tk.Frame(tabs_content_frame, bg='white')
    
    # Crear canvas y scrollbar para contenido scrolleable
    datalogger_canvas = tk.Canvas(datalogger_content_frame, bg='white', highlightthickness=0, bd=0)
    datalogger_scrollbar = ttk.Scrollbar(datalogger_content_frame, orient="vertical", command=datalogger_canvas.yview)
    
    # Frame centrador para el contenido
    datalogger_centrador = tk.Frame(datalogger_canvas, bg='white')
    datalogger_scrollable_frame = tk.Frame(datalogger_centrador, bg='white')
    datalogger_scrollable_frame.pack(anchor='center', expand=True)

    # Centrar el contenido horizontalmente al redimensionar
    def resize_datalogger_canvas(event):
        canvas_width = event.width
        datalogger_centrador.config(width=canvas_width)
        datalogger_canvas.itemconfig(datalogger_window_id, width=canvas_width)
    datalogger_canvas.bind('<Configure>', resize_datalogger_canvas)

    datalogger_centrador.pack(expand=True)
    datalogger_window_id = datalogger_canvas.create_window((0, 0), window=datalogger_centrador, anchor="n")
    datalogger_canvas.configure(yscrollcommand=datalogger_scrollbar.set)

    datalogger_scrollable_frame.bind(
        "<Configure>",
        lambda e: datalogger_canvas.configure(scrollregion=datalogger_canvas.bbox("all"))
    )
    
    # Configurar scroll con mouse wheel
    def _on_mousewheel(event):
        datalogger_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    
    datalogger_canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    # Desvincular mouse wheel cuando se salga del canvas
    def _on_leave(event):
        datalogger_canvas.unbind_all("<MouseWheel>")
    
    def _on_enter(event):
        datalogger_canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    datalogger_canvas.bind('<Enter>', _on_enter)
    datalogger_canvas.bind('<Leave>', _on_leave)

    # Empaquetar canvas y scrollbar
    datalogger_canvas.pack(side="left", fill="both", expand=True, padx=(0, 5))
    datalogger_scrollbar.pack(side="right", fill="y")
    
    # Título de datalogger
    datalogger_titulo = tk.Label(datalogger_scrollable_frame, text='Datalogger:', 
                                font=('Segoe UI', 10, 'bold'), 
                                fg='#25364a', bg='white')
    datalogger_titulo.pack(anchor='w', padx=10, pady=(8, 5))
    
    # Frame para mostrar ruta del archivo Excel
    ruta_frame = tk.Frame(datalogger_scrollable_frame, bg='#e8f4fd', relief='solid', bd=1)
    ruta_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Título de ruta
    ruta_titulo = tk.Label(ruta_frame, text='Archivo Excel:', 
                          font=('Segoe UI', 8, 'bold'), 
                          fg='#25364a', bg='#e8f4fd')
    ruta_titulo.pack(anchor='w', padx=10, pady=(5, 2))
    
    # Label para mostrar ruta del archivo
    ruta_label = tk.Label(ruta_frame, text='No se ha configurado el dispositivo', 
                          font=('Segoe UI', 8), 
                          fg='#e74c3c', bg='#e8f4fd',
                          justify='left', anchor='nw')
    ruta_label.pack(anchor='w', padx=10, pady=(0, 5))
    
    # Frame para controles del datalogger
    datalogger_controls_frame = tk.Frame(datalogger_scrollable_frame, bg='white')
    datalogger_controls_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Variable para el estado del datalogger
    datalogger_running = tk.BooleanVar(value=False)
    
    # Variables para almacenar datos de la gráfica
    frecuencias_grafica = []
    tiempos_relativos_grafica = []
    
    # Crear figura de matplotlib para la gráfica
    fig, ax = plt.subplots(figsize=(8, 4))
    ax.set_xlabel('Tiempo Relativo (s)', fontsize=10)
    ax.set_ylabel('Frecuencia (Hz)', fontsize=10)
    ax.set_title('Frecuencia vs Tiempo Relativo', fontsize=12, fontweight='bold')
    ax.grid(True, alpha=0.3)
    
    # Crear canvas de matplotlib
    canvas_frame = tk.Frame(datalogger_scrollable_frame, bg='white')
    canvas_frame.pack(fill='both', expand=True, padx=10, pady=(0, 10))
    
    canvas = FigureCanvasTkAgg(fig, canvas_frame)
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.pack(fill='both', expand=True)
    
    # Frame para estadísticas
    stats_frame = tk.Frame(datalogger_scrollable_frame, bg='#f8f9fa', relief='solid', bd=1)
    stats_frame.pack(fill='x', padx=10, pady=(0, 10))
    
    # Título de estadísticas
    stats_titulo = tk.Label(stats_frame, text='Estadísticas en Tiempo Real:', 
                           font=('Segoe UI', 9, 'bold'), 
                           fg='#25364a', bg='#f8f9fa')
    stats_titulo.pack(anchor='w', padx=10, pady=(8, 5))
    
    # Label para mostrar estadísticas
    stats_label = tk.Label(stats_frame, text='', 
                          font=('Segoe UI', 8), 
                          fg='#2c3e50', bg='#f8f9fa',
                          justify='left', anchor='nw')
    stats_label.pack(anchor='w', padx=10, pady=(0, 8))
    
    # Función para actualizar la gráfica
    def actualizar_grafica():
        if len(frecuencias_grafica) > 0:
            # Limpiar gráfica anterior
            ax.clear()
            
            # Crear nueva gráfica
            ax.plot(tiempos_relativos_grafica, frecuencias_grafica, 'b-', linewidth=1, alpha=0.8)
            ax.scatter(tiempos_relativos_grafica, frecuencias_grafica, c='red', s=20, alpha=0.6)
            
            # Configurar ejes
            ax.set_xlabel('Tiempo Relativo (s)', fontsize=10)
            ax.set_ylabel('Frecuencia (Hz)', fontsize=10)
            ax.set_title('Frecuencia vs Tiempo Relativo', fontsize=12, fontweight='bold')
            ax.grid(True, alpha=0.3)
            
            # Auto-escalado inteligente
            if len(tiempos_relativos_grafica) > 1:
                # Para el eje X: mostrar los últimos 50 puntos o todos si hay menos
                x_max = max(tiempos_relativos_grafica)
                x_min = max(0, x_max - 50 * (x_max / len(tiempos_relativos_grafica)))
                ax.set_xlim(x_min, x_max)
                
                # Para el eje Y: margen del 5% arriba y abajo
                y_min, y_max = min(frecuencias_grafica), max(frecuencias_grafica)
                y_range = y_max - y_min
                if y_range > 0:
                    margin = y_range * 0.05
                    ax.set_ylim(y_min - margin, y_max + margin)
            
            # Actualizar canvas
            canvas.draw()
    
    # Función para actualizar estadísticas
    def actualizar_estadisticas():
        if len(frecuencias_grafica) > 0:
            freqs_array = np.array(frecuencias_grafica)
            
            # Calcular estadísticas
            maximo = np.max(freqs_array)
            minimo = np.min(freqs_array)
            media = np.mean(freqs_array)
            mediana = np.median(freqs_array)
            desv_tipica = np.std(freqs_array)
            varianza = np.var(freqs_array)
            
            # Formatear estadísticas
            stats_text = f"""Máximo: {maximo:.6f} Hz
Mínimo: {minimo:.6f} Hz
Media: {media:.6f} Hz
Mediana: {mediana:.6f} Hz
Desv. Típica: {desv_tipica:.6f} Hz
Varianza: {varianza:.6f} Hz²
Nº Muestras: {len(frecuencias_grafica)}"""
            
            stats_label.config(text=stats_text)
    
    # Función para agregar punto a la gráfica
    def agregar_punto_grafica(frecuencia, tiempo_relativo):
        frecuencias_grafica.append(frecuencia)
        tiempos_relativos_grafica.append(tiempo_relativo)
        
        # Actualizar gráfica y estadísticas
        actualizar_grafica()
        actualizar_estadisticas()
    
    # Función para finalizar medición definitivamente
    def finalizar_medicion():
        global medicion_pausada
        
        try:
            # Guardar estadísticas finales en Excel si hay datos
            if len(frecuencias_grafica) > 0:
                try:
                    # Calcular estadísticas finales
                    freqs_array = np.array(frecuencias_grafica)
                    maximo = np.max(freqs_array)
                    minimo = np.min(freqs_array)
                    media = np.mean(freqs_array)
                    mediana = np.median(freqs_array)
                    desv_tipica = np.std(freqs_array)
                    varianza = np.var(freqs_array)
                    
                    # Guardar estadísticas en una nueva hoja del Excel
                    from openpyxl import load_workbook
                    wb = load_workbook(cnt_device.file_path)
                    
                    # Crear hoja de estadísticas
                    if 'Estadísticas' in wb.sheetnames:
                        ws_stats = wb['Estadísticas']
                    else:
                        ws_stats = wb.create_sheet('Estadísticas')
                    
                    # Limpiar hoja de estadísticas
                    ws_stats.delete_rows(1, ws_stats.max_row)
                    
                    # Escribir estadísticas
                    ws_stats.append(['ESTADÍSTICAS FINALES DE LA MEDICIÓN'])
                    ws_stats.append([])
                    ws_stats.append(['Parámetro', 'Valor', 'Unidad'])
                    ws_stats.append(['Máximo', maximo, 'Hz'])
                    ws_stats.append(['Mínimo', minimo, 'Hz'])
                    ws_stats.append(['Media', media, 'Hz'])
                    ws_stats.append(['Mediana', mediana, 'Hz'])
                    ws_stats.append(['Desviación Típica', desv_tipica, 'Hz'])
                    ws_stats.append(['Varianza', varianza, 'Hz²'])
                    ws_stats.append(['Número de Muestras', len(frecuencias_grafica), ''])
                    
                    # Guardar Excel
                    wb.save(cnt_device.file_path)
                    wb.close()
                    
                    print("Estadísticas finales guardadas en Excel.")
                    
                except Exception as e:
                    print(f"Advertencia al guardar estadísticas: {e}")
            
            # Cerrar el archivo Excel de manera definitiva
            try:
                cnt_device.cerrar_archivo_excel()
                print("Archivo Excel cerrado definitivamente.")
            except Exception as e:
                print(f"Advertencia al cerrar archivo Excel: {e}")
            
            # Abortar la medición continua si está activa
            try:
                cnt_device.abort_continuous_measurement()
            except Exception as e:
                print(f"Advertencia al abortar medición: {e}")
            
            # Resetear estados
            datalogger_running.set(False)
            medicion_pausada = False
            
            # Limpiar ruta del archivo Excel
            ruta_archivo_excel = None
            ruta_label.config(text='Archivo Excel cerrado', fg='#e74c3c')
            
            # Restaurar estado de la interfaz
            btn_start_stop.config(text='▶️  Iniciar Datalogger', bg='#27ae60', fg='white', cursor='hand2', state='normal')
            btn_fin_medicion.config(state='disabled')
            status_label.config(text='Estado: Medición finalizada', fg='#6c757d')
            
            # Actualizar información de configuración
            actualizar_info_configuracion()
            
            tk.messagebox.showinfo('Medición Finalizada', 'La medición se ha finalizado definitivamente. Las estadísticas han sido guardadas en el archivo Excel.')
            
        except Exception as e:
            tk.messagebox.showerror('Error', f'Error al finalizar medición: {str(e)}')
    
    # Función para iniciar/detener datalogger
    def toggle_datalogger():
        global cnt_device, medicion_pausada
        
        if not datalogger_running.get():
            # Verificar que la configuración esté guardada
            if not configuracion_guardada:
                tk.messagebox.showerror('Error', 'Debe guardar la configuración antes de iniciar el datalogger.')
                return
            
            # Verificar que el dispositivo esté conectado
            if cnt_device is None:
                tk.messagebox.showerror('Error', 'Dispositivo no conectado. Conecte el CNT-91 primero.')
                return
            
            try:
                # Si es una nueva medición (no pausada), limpiar datos
                if not medicion_pausada:
                    # Limpiar datos de gráfica anteriores
                    frecuencias_grafica.clear()
                    tiempos_relativos_grafica.clear()
                    actualizar_grafica()
                    actualizar_estadisticas()
                    
                    # Configurar el dispositivo usando las variables globales
                    impedancia_convertida = 'MAX' if impedancia == 'Max' else 'MIN'
                    filtro_analog_bool = True if filtro_Analog_PASSAbaja == 'True' else False
                    
                    # Configurar el dispositivo
                    file_path = cnt_device.configurar_dispositivo(
                        canal=canal_seleccionado,
                        intervalo_s=intervalo_s,
                        acoplamiento=acoplamiento,
                        impedancia=impedancia_convertida,
                        atenuacion=atenuacion,
                        trigger_level=trigger_level,
                        trigger_slope=trigger_slope,
                        filtro_Digital_PASSAbaja=None,
                        filtro_Analog_PASSAbaja=filtro_analog_bool,
                        file_path=None
                    )
                    
                    # Actualizar la ruta del archivo Excel
                    ruta_archivo_excel = file_path
                    ruta_label.config(text=f'Guardando en: {file_path}', fg='#27ae60')
                    
                    # Iniciar medición continua
                    tiempo_espera = cnt_device.start_continuous_measurement(
                        intervalo_s=intervalo_s, 
                        canal=canal_seleccionado
                    )
                else:
                    # Si es reanudación, solo reiniciar la medición continua
                    tiempo_espera = cnt_device.start_continuous_measurement(
                        intervalo_s=intervalo_s, 
                        canal=canal_seleccionado
                    )
                
                # Calcular parámetros de buffer según el intervalo
                if intervalo_s < 2:
                    tiempo_espera = 0
                    lenght = 10
                elif intervalo_s < 5:
                    lenght = 1
                    tiempo_espera = 2.5 * (intervalo_s - 2) ** 2 + 0.09
                elif intervalo_s < 10:
                    lenght = 2
                    tiempo_espera = 1.2 * intervalo_s + 0.09
                else:
                    lenght = 1
                    tiempo_espera = intervalo_s + 5
                
                # Variables para el bucle de medición
                buffer_frecs = []
                t0 = None  # Primer timestamp para calcular tiempos relativos
                
                # Si es reanudación, usar el último tiempo relativo como base
                if medicion_pausada and len(tiempos_relativos_grafica) > 0:
                    t0_offset = max(tiempos_relativos_grafica)
                else:
                    t0_offset = 0
                
                # Función interna para el bucle de medición
                def medicion_loop():
                    nonlocal buffer_frecs, t0
                    
                    if datalogger_running.get():
                        try:
                            # Obtener muestras
                            frecs, ts = cnt_device.fetch_continuous_samples(
                                n_muestras=1,
                                tiempo_espera=tiempo_espera
                            )
                            
                            for f, t in zip(frecs, ts):
                                if t0 is None:
                                    t0 = t
                                t_rel = t - t0 + t0_offset
                                buffer_frecs.append((f, t, t_rel))
                                
                                # Agregar punto a la gráfica en tiempo real
                                agregar_punto_grafica(f, t_rel)
                                
                                # Actualizar estado en la interfaz
                                status_label.config(text=f'Estado: Midiendo... Frecuencia: {f:.6f} Hz, T. Relativo: {t_rel:.3f} s', fg='#27ae60')
                                
                                # Actualizar información de configuración con datos en tiempo real
                                info_text = f"""Canal: {canal_seleccionado}
Intervalo: {intervalo_s:.6f} s
Acoplamiento: {acoplamiento}
Impedancia: {impedancia}
Atenuación: {atenuacion}x
Trigger: {'Automático' if trigger_level is None else f'{trigger_level:.1f}V'}
Pendiente: {trigger_slope}
Filtro Analógico: {'True' if filtro_Analog_PASSAbaja == 'True' else 'False'}

Última medición:
Frecuencia: {f:.6f} Hz
Timestamp: {t:.6f} s
T. Relativo: {t_rel:.6f} s"""
                                info_label.config(text=info_text)
                            
                            # Guardar en Excel cuando el buffer esté lleno
                            if len(buffer_frecs) >= lenght:
                                for f, t, t_rel in buffer_frecs:
                                    cnt_device.append_measurement(f, t, t_rel)
                                buffer_frecs.clear()
                            
                            # Programar la siguiente medición
                            root.after(100, medicion_loop)  # 100ms entre mediciones
                            
                        except Exception as e:
                            tk.messagebox.showerror('Error', f'Error durante la medición: {str(e)}')
                            datalogger_running.set(False)
                            btn_start_stop.config(text='▶️  Iniciar Datalogger', bg='#27ae60', fg='white', cursor='hand2')
                            status_label.config(text='Estado: Error en medición', fg='#e74c3c')
                
                # Iniciar el bucle de medición
                datalogger_running.set(True)
                medicion_pausada = False
                btn_start_stop.config(text='⏸️  Pausar Datalogger', bg='#f39c12', fg='white', cursor='hand2')
                btn_fin_medicion.config(state='normal')
                status_label.config(text='Estado: Iniciando medición...', fg='#27ae60')
                
                # Iniciar el bucle de medición
                medicion_loop()
                
            except Exception as e:
                tk.messagebox.showerror('Error', f'Error al iniciar datalogger: {str(e)}')
                datalogger_running.set(False)
                btn_start_stop.config(text='▶️  Iniciar Datalogger', bg='#27ae60', fg='white', cursor='hand2')
                status_label.config(text='Estado: Error al iniciar', fg='#e74c3c')
        else:
            # Pausar datalogger (no cerrar Excel)
            try:
                datalogger_running.set(False)
                medicion_pausada = True
                
                # Solo abortar la medición continua, NO cerrar Excel
                try:
                    cnt_device.abort_continuous_measurement()
                except Exception as e:
                    print(f"Advertencia al abortar medición: {e}")
                
                # Cambiar botón a "Reanudar"
                btn_start_stop.config(text='▶️  Reanudar Datalogger', bg='#27ae60', fg='white', cursor='hand2')
                status_label.config(text='Estado: Datalogger pausado', fg='#f39c12')
                
                tk.messagebox.showinfo('Datalogger Pausado', 'La medición se ha pausado. Puede reanudarla o finalizarla definitivamente.')
                
            except Exception as e:
                tk.messagebox.showerror('Error', f'Error al pausar datalogger: {str(e)}')
    
    # Botón para iniciar/detener datalogger
    btn_start_stop = tk.Button(datalogger_controls_frame, text='▶️  Iniciar Datalogger', 
                              command=toggle_datalogger, 
                              font=('Segoe UI', 10, 'bold'),
                              bg='#cccccc', fg='#666666',
                              relief='flat', padx=15, pady=5,
                              cursor='arrow',
                              state='disabled')
    btn_start_stop.pack(side='left', pady=(3, 0))
    
    # Botón para finalizar medición (rojo)
    btn_fin_medicion = tk.Button(datalogger_controls_frame, text='🔴  FIN DE MEDICIÓN', 
                                command=lambda: finalizar_medicion(),
                                font=('Segoe UI', 10, 'bold'),
                                bg='#e74c3c', fg='white',
                                relief='flat', padx=15, pady=5,
                                cursor='hand2',
                                state='disabled')
    btn_fin_medicion.pack(side='left', padx=(10, 0), pady=(3, 0))
    
    # Label para mostrar estado del datalogger
    status_label = tk.Label(datalogger_controls_frame, text='Estado: Configure el dispositivo primero', 
                           font=('Segoe UI', 9), 
                           fg='#e74c3c', bg='white')
    status_label.pack(side='left', padx=(15, 0), pady=(3, 0))
    
    # Separador
    separador_datalogger = tk.Frame(datalogger_scrollable_frame, height=1, bg='#e0e0e0')
    separador_datalogger.pack(fill='x', padx=10, pady=10)
    
    # Frame para información de configuración actual
    info_frame = tk.Frame(datalogger_scrollable_frame, bg='#f8f9fa', relief='solid', bd=1)
    info_frame.pack(fill='x', padx=10, pady=(0, 10))
    
    # Título de información
    info_titulo = tk.Label(info_frame, text='Configuración Actual:', 
                          font=('Segoe UI', 9, 'bold'), 
                          fg='#25364a', bg='#f8f9fa')
    info_titulo.pack(anchor='w', padx=10, pady=(8, 5))
    
    # Label para mostrar información de configuración
    info_label = tk.Label(info_frame, text='', 
                         font=('Segoe UI', 8), 
                         fg='#2c3e50', bg='#f8f9fa',
                         justify='left', anchor='nw')
    info_label.pack(anchor='w', padx=10, pady=(0, 8))
    
    # Función para actualizar información de configuración
    def actualizar_info_configuracion():
        # Formatear valores para mostrar
        if intervalo_s < 0.001:
            intervalo_formato = f"{intervalo_s:.2e}"
        elif intervalo_s < 1:
            intervalo_formato = f"{intervalo_s:.6f}"
        else:
            intervalo_formato = f"{intervalo_s:.3f}"
        
        trigger_formato = "Automático" if trigger_level is None else f"{trigger_level:.1f}V"
        filtro_formato = "True" if filtro_Analog_PASSAbaja == 'True' else "False"
        
        info_text = f"""Canal: {canal_seleccionado}
Intervalo: {intervalo_formato} s
Acoplamiento: {acoplamiento}
Impedancia: {impedancia}
Atenuación: {atenuacion}x
Trigger: {trigger_formato}
Pendiente: {trigger_slope}
Filtro Analógico: {filtro_formato}"""
        
        info_label.config(text=info_text)
    
    # Mostrar la pestaña de configuración por defecto
    switch_tab('config')

# Función para mostrar la página de Allan Deviation vs tau
def mostrar_allan_deviation(widgets):
    frame_contenido = widgets['frame_contenido']
    
    # Limpiar el frame de contenido
    for widget in frame_contenido.winfo_children():
        widget.destroy()
    
    # Frame principal con padding
    main_frame = tk.Frame(frame_contenido, bg='#f6f7fa')
    main_frame.pack(fill='both', expand=True, padx=10, pady=5)
    
    # Título principal - Allan Deviation vs tau
    titulo_frame = tk.Frame(main_frame, bg='#f6f7fa')
    titulo_frame.pack(fill='x', pady=(0, 10))
    
    titulo = tk.Label(titulo_frame, text='Allan Deviation vs tau', 
                     font=('Segoe UI', 16, 'bold'), 
                     fg='#25364a', bg='#f6f7fa')
    titulo.pack(anchor='w')
    
    # Línea separadora azul
    separador = tk.Frame(titulo_frame, height=2, bg='#2980f2')
    separador.pack(fill='x', pady=(5, 0))
    
    # Frame de contenido principal
    contenido_frame = tk.Frame(main_frame, bg='white', relief='flat', bd=1)
    contenido_frame.pack(fill='both', expand=True, padx=0, pady=10)
    
    # Mensaje temporal
    mensaje = tk.Label(contenido_frame, text='Página en desarrollo...', 
                      font=('Segoe UI', 12), 
                      fg='#666666', bg='white')
    mensaje.pack(expand=True)

# Función para mostrar la página de información del CNT-91
def mostrar_informacion_cnt91(widgets):
    frame_contenido = widgets['frame_contenido']
    
    # Limpiar el frame de contenido
    for widget in frame_contenido.winfo_children():
        widget.destroy()
    
    # Frame principal con padding
    main_frame = tk.Frame(frame_contenido, bg='#f6f7fa')
    main_frame.pack(fill='both', expand=True, padx=10, pady=5)
    
    # Título principal - Información CNT-91
    titulo_frame = tk.Frame(main_frame, bg='#f6f7fa')
    titulo_frame.pack(fill='x', pady=(0, 10))
    
    titulo = tk.Label(titulo_frame, text='Información CNT-91', 
                     font=('Segoe UI', 16, 'bold'), 
                     fg='#25364a', bg='#f6f7fa')
    titulo.pack(anchor='w')
    
    # Línea separadora
    separador = tk.Frame(titulo_frame, height=2, bg='#2980f2')
    separador.pack(fill='x', pady=(5, 0))
    
    # Frame de contenido con scroll
    contenido_frame = tk.Frame(main_frame, bg='white', relief='solid', bd=1)
    contenido_frame.pack(fill='both', expand=True, pady=(0, 5))
    
    # Canvas y scrollbar para contenido scrolleable
    canvas = tk.Canvas(contenido_frame, bg='white', highlightthickness=0, bd=0)
    scrollbar = ttk.Scrollbar(contenido_frame, orient="vertical", command=canvas.yview)
    # Frame centrador para el contenido
    centrador = tk.Frame(canvas, bg='white')
    scrollable_frame = tk.Frame(centrador, bg='white')
    scrollable_frame.pack(anchor='center', expand=True)

    # Centrar el contenido horizontalmente al redimensionar
    def resize_canvas(event):
        canvas_width = event.width
        centrador.config(width=canvas_width)
        canvas.itemconfig(window_id, width=canvas_width)
    canvas.bind('<Configure>', resize_canvas)

    centrador.pack(expand=True)
    window_id = canvas.create_window((0, 0), window=centrador, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    # Empaquetar canvas y scrollbar con mejor espaciado
    canvas.pack(side="left", fill="both", expand=True, padx=(10, 5), pady=10)
    scrollbar.pack(side="right", fill="y", padx=(0, 10), pady=10)
    
    # Contenido de la información
    contenido = get_info_cnt91_sections()
    
    # Crear el contenido con formato profesional
    for i, (titulo_seccion, texto) in enumerate(contenido):
        # Título de sección
        titulo_label = tk.Label(scrollable_frame, text=titulo_seccion,
                               font=('Segoe UI', 12, 'bold'),
                               fg='#25364a', bg='white',
                               anchor='w', justify='left')
        titulo_label.pack(fill='x', padx=15, pady=(20 if i == 0 else 15, 5))
        
        # Texto de la sección
        texto_label = tk.Label(scrollable_frame, text=texto,
                              font=('Segoe UI', 10),
                              fg='#2c3e50', bg='white',
                              anchor='nw', justify='left',
                              wraplength=800)
        texto_label.pack(fill='x', padx=15, pady=(0, 10))
        
        # Separador entre secciones (excepto la última)
        if i < len(contenido) - 1:
            separador_seccion = tk.Frame(scrollable_frame, height=1, bg='#e0e0e0')
            separador_seccion.pack(fill='x', padx=15, pady=5)
    
    # --- Sección de Documentación y Recursos ---
    recursos_frame = tk.Frame(scrollable_frame, bg='white')
    recursos_frame.pack(fill='x', padx=15, pady=(30, 10), anchor='w')

    # Título de la sección
    recursos_titulo = tk.Label(recursos_frame, text='Documentación y Recursos',
                               font=('Segoe UI', 13, 'bold'), fg='#25364a', bg='white', anchor='w')
    recursos_titulo.pack(anchor='w', pady=(0, 10))

    recursos = get_info_cnt91_resources()

    def abrir_url(url):
        import webbrowser
        webbrowser.open_new(url)

    for icono, texto, url in recursos:
        frame = tk.Frame(recursos_frame, bg='white')
        frame.pack(fill='x', pady=4, anchor='w')
        label_icon = tk.Label(frame, text=icono, font=('Segoe UI Symbol', 13), fg='#25364a', bg='white')
        label_icon.pack(side='left', padx=(0, 8))
        enlace = tk.Label(frame, text=texto, font=('Segoe UI', 11, 'bold'), fg='#25364a', bg='white', cursor='hand2')
        enlace.pack(side='left')
        # Efecto hover
        def on_enter(e, l=enlace):
            l.config(fg='#2980f2', underline=True)
        def on_leave(e, l=enlace):
            l.config(fg='#25364a', underline=False)
        enlace.bind('<Enter>', on_enter)
        enlace.bind('<Leave>', on_leave)
        enlace.bind('<Button-1>', lambda e, url=url: abrir_url(url))
    
    # Configurar el scroll para que funcione correctamente
    canvas.update_idletasks()
    canvas.configure(scrollregion=canvas.bbox("all"))

# Función para manejar la conexión al dispositivo
def conectar_dispositivo(widgets):
    global cnt_device
    entry_id = widgets['entry_id']
    estado_label = widgets['estado']
    btn_conectar = widgets['btn_conectar']
    btn_mediciones = widgets['btn_mediciones']
    address = entry_id.get().strip()
    # Si el campo está vacío, usar el valor por defecto
    if not address:
        address = DEFAULT_GPIB
        entry_id.delete(0, tk.END)
        entry_id.insert(0, DEFAULT_GPIB)
    # Cambiar estado a "Trying to connect" en naranja
    estado_label.config(text='Estado:  Trying to connect', fg='#e67e22')
    widgets['frame_superior'].update_idletasks()
    try:
        cnt_device = CNT.CNT_frequenciometro(address)
        # Si no hay excepción, conexión exitosa
        estado_label.config(text='Estado:  Conectado', fg='#27ae60')
        # Cambiar el botón a rojo con texto "Desconectar"
        btn_conectar.config(text='🔌  Desconectar', style='DangerSidebar.TButton')
        # Cambiar la función del botón para desconectar
        btn_conectar.config(command=lambda: desconectar_dispositivo(widgets))
        # Actualizar el estado del botón
        btn_conectar.estado_conectado = True
        # Habilitar botón Mediciones
        btn_mediciones.config(state='normal')
    except Exception as e:
        estado_label.config(text='Estado:  Desconectado', fg='#e74c3c')
        tk.messagebox.showerror('Error de conexión', 'No se logró conexión con el Dispositivo.\nRevise alimentación o instalación de drivers de comunicación.')
        btn_mediciones.config(state='disabled')

# Función para manejar la desconexión del dispositivo
def desconectar_dispositivo(widgets):
    global cnt_device
    estado_label = widgets['estado']
    btn_conectar = widgets['btn_conectar']
    btn_mediciones = widgets['btn_mediciones']
    
    try:
        if cnt_device is not None:
            # Cerrar la conexión usando la función que creamos
            cnt_device.cerrar_conexion()
            cnt_device = None
            # Cambiar estado a desconectado
            estado_label.config(text='Estado:  Desconectado', fg='#e74c3c')
            # Cambiar el botón de vuelta a azul con texto "Conectar"
            btn_conectar.config(text='🔌  Conectar', style='PrimarySidebar.TButton')
            # Cambiar la función del botón para conectar
            btn_conectar.config(command=lambda: conectar_dispositivo(widgets))
            # Actualizar el estado del botón
            btn_conectar.estado_conectado = False
            btn_mediciones.config(state='disabled')
            print("Dispositivo desconectado correctamente.")
    except Exception as e:
        print(f"Error al desconectar: {e}")
        # Aún así, resetear el estado de la interfaz
        estado_label.config(text='Estado:  Desconectado', fg='#e74c3c')
        btn_conectar.config(text='🔌  Conectar', style='PrimarySidebar.TButton')
        btn_conectar.config(command=lambda: conectar_dispositivo(widgets))
        btn_conectar.estado_conectado = False
        btn_mediciones.config(state='disabled')

# Función para manejar el cierre de la ventana
def on_closing(root, widgets):
    """Maneja el cierre correcto de la aplicación"""
    try:
        # Si hay un dispositivo conectado, desconectarlo
        if cnt_device is not None:
            print("Cerrando conexión con el dispositivo...")
            desconectar_dispositivo(widgets)
        
        # Si hay una medición en curso, finalizarla
        if 'datalogger_running' in globals() and datalogger_running.get():
            print("Finalizando medición en curso...")
            try:
                # Abortar medición continua
                cnt_device.abort_continuous_measurement()
            except:
                pass
        
        # Si hay un archivo Excel abierto, cerrarlo
        if 'ruta_archivo_excel' in globals() and ruta_archivo_excel is not None:
            print("Cerrando archivo Excel...")
            try:
                cnt_device.cerrar_archivo_excel()
            except:
                pass
        
        print("Cerrando aplicación...")
        
    except Exception as e:
        print(f"Error durante el cierre: {e}")
    
    finally:
        # Destruir la ventana principal
        root.destroy()
        # Salir del programa
        import sys
        sys.exit(0)

if __name__ == '__main__':
    # Crear la ventana principal de la aplicación
    root = tk.Tk()
    widgets = crear_layout_principal(root)

    # Configurar el protocolo de cierre de ventana
    root.protocol("WM_DELETE_WINDOW", lambda: on_closing(root, widgets))

    # Importar messagebox aquí para evitar problemas de importación circular
    import tkinter.messagebox
    tk.messagebox = tkinter.messagebox

    # Deshabilitar botón Mediciones por defecto
    widgets['btn_mediciones'].config(state='disabled')

    # Asignar la función al botón de conectar
    widgets['btn_conectar'].config(command=lambda: conectar_dispositivo(widgets))

    # Asignar función al botón Allan Deviation vs tau
    widgets['btn_config'].config(command=lambda: mostrar_allan_deviation(widgets))

    # Asignar función al botón Mediciones
    widgets['btn_mediciones'].config(command=lambda: mostrar_menu_canal(widgets))

    # Asignar función al botón Información CNT-91
    widgets['btn_info'].config(command=lambda: mostrar_informacion_cnt91(widgets))

    # Iniciar el bucle principal de la interfaz gráfica (espera eventos del usuario)
    root.mainloop() 