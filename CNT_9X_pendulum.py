"""

Created on Mon Feb 24 11:47 2025

@author: Carlos Darvoy Espigulé

Objetivo : crear clase de un instrumento de medida

MI MODELO : Timer/Counter/Analyzer CNT_91

"""
""" Comunicación PVISA;  """

    
""" ═━═━═━═━═━═━═━═━═━═━═━═ Definición de Clase ━═━═━═━═━═━═━═━━═━═━═━═━ """
import numpy as np
import time
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import datetime
from openpyxl.worksheet.table import Table, TableStyleInfo
import atexit

class CNT_frequenciometro: 
   
   
    """ Address: Dirección del recurso VISA 
    Por defecto -> "GPIB0::18::INSTR" instrumento conectado vía GPIB """   
    def __init__(self, address ='GPIB0::10::INSTR'):
        self.address = address
        self.dev = None
        self._conexion_cerrada = False
        self.file_path = None
        
        # Registrar función de limpieza automática
        atexit.register(self._cleanup_on_exit)
        
        # Configurar conexión
        self._setup_connection()
    
    def _cleanup_on_exit(self):
        """
        Función de limpieza automática que se ejecuta al finalizar el programa.
        Asegura que los recursos se liberen correctamente.
        """
        try:
            if not self._conexion_cerrada:
                self.cerrar_conexion()
        except:
            pass  # En atexit, no queremos que se lancen excepciones
    
    def _setup_connection(self):
        import pyvisa
        from pyvisa.highlevel import ResourceManager
        
        #Detección y Administración de conexion con el instrumento
        rm = ResourceManager()
        self.dev=rm.open_resource(self.address)
        #Utiliza el objeto rm para abrir una 
        #conexión con el instrumento especificado por la dirección address.
        #self.dev: representa el instrumento para enviar,recibir 
        #self.dev.write('*CLS')  # Limpiar errores previos
        self.dev.write('*IDN?')
        #Envio comando SCPI  -> *IDN? (identificación estándar)
        #permite verificar que la comunicación se ha establecido correctamente.
        
        resposta=self.dev.read()
        #Lee la respuesta que el instrumento envía tras recibir el comando *IDN?.
        #almacenada en la variable resposta, generalmente contiene información 
        #identificativa del dispositivo.
        
        self.dev.write('*OPT?')
        #Envio comando SCPI  -> *IDN? (identificación estándar)
        #permite verificar que la comunicación se ha establecido correctamente.
        
        resposta2=self.dev.read()
        #Lee 
        
        print('Communication established with GSE-UIB ' + resposta  )
        #Enseña la respuesta          
        print('Options instaled   ' + resposta2  )

        

   

   # UNA MEDICIÓN DE FREQUENCIA
    def measure_frequency(self, channel='A'):
        
        """
        Realiza una medición única de frecuencia en el canal especificado.
        
        Parámetros:
          channel (str o int): Puede ser 'A', 'B', 1 o 2.
                               - 'A' o 1 medirá el canal A (comando MEASure:FREQ? (@1))
                               - 'B' o 2 medirá el canal B (comando MEASure:FREQ? (@2))
                               Por defecto se utiliza el canal A.
        
        Retorna:
          La respuesta del instrumento con el valor medido.
        """
        # Limpieza de errores previos
        self.dev.write("*CLS")
        
        # Diccionario para mapear la entrada a su comando correspondiente
        canales = {'A': '@1', 'B': '@2', '1': '@1', '2': '@2'}
        
        canal_seleccionado = str(channel).upper()
        if canal_seleccionado not in canales:
            raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
        
        # Construir y enviar el comando basado en el canal seleccionado
        comando = f"MEASure:FREQ? ({canales[canal_seleccionado]})"
        self.dev.write(comando)
        
        # Leer y retornar la respuesta del instrumento
        response = self.dev.read()
        return response
    
    
    
    








# MEDICION DE TEMPERATURA
        
        """  Es La temperatura a la salida del fan del instrumento en el controlador """   
    def Measure_temperature_example(self):
             
             self.dev.write(':SYST:TEMP?')
             temp = self.dev.read()
             return temp 






# JOAN 2

# MEDICIÓN CONTINUA DE FREQUENCIAS [[Medición continua]+[FETCH on FLY]]
# Especialmente rápida en solo 50s ha sacado 1700 medidas, Pero el tiempo que tarda es desconocido

    def measure_frequency_array_CONTINUOUS(self, duration_s, channel='A'):
        """
        Mide frecuencias de manera continua durante un tiempo dado, usando el canal especificado.
    
        Parámetros:
            duration_s (float): Duración total de la medida en segundos.
            channel (str|int): Canal a medir ('A', 'B', 1 o 2). Por defecto 'A'.
    
        Retorna:
            lista de floats con los valores medidos.
        """
        import time
    
        # ========== SECCIÓN 1: Selección y validación del canal ==========
        # Diccionario para mapear la selección de canal del usuario a la sintaxis SCPI correcta
        canales = {'A': '@1', 'B': '@2', '1': '@1', '2': '@2'}
        ch = str(channel).upper()
        if ch not in canales:
            raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
        canal_cmd = canales[ch]
    
        # ========== SECCIÓN 2: Configuración inicial del instrumento ==========
        # 1. Resetear el instrumento para asegurar estado limpio
        self.dev.write("*RST")
        # 2. Limpiar errores previos para evitar problemas durante la medición
        self.dev.write("*CLS")
        # 3. Desactivar cálculos internos automáticos que podrían interferir con la adquisición continua
        self.dev.write(":CALC:STAT OFF")
        # 4. Configurar el instrumento para medición continua de frecuencia en el canal elegido
        self.dev.write(f":CONF:ARR:FREQ? ( ,({canal_cmd}))")
        # 5. Iniciar la adquisición continua
        self.dev.write(":INIT:CONT ON")
    
        # ========== SECCIÓN 3: Adquisición continua de datos ==========
        t0 = time.time()         # Guardar el instante de inicio
        results = []             # Lista para almacenar resultados
        try:
            while (time.time() - t0) < duration_s:
                # 5. Solicitar el último valor disponible sin detener la adquisición continua
                self.dev.write("FETC:ARR? -1")
                resp = self.dev.read().strip()
                try:
                    # 6. Intentar convertir la respuesta a float y almacenar
                    val = float(resp)
                    results.append(val)
                except ValueError:
                    # 7. Si la respuesta no es un número, ignorar (pueden aparecer respuestas vacías o errores de comunicación)
                    pass
        finally:
            # ========== SECCIÓN 4: Finalización segura ==========
            # 8. Asegurarse de detener la adquisición continua al final, ocurra lo que ocurra en el bucle
            self.dev.write(":INIT:CONT OFF")
    
        # ========== SECCIÓN 5: Devolver resultados ==========
        return results
    
    


#JOAN 3 PROBAMOS OTRO METODO Usar el modo de medición con "Sample Timer"
# JOAN 3
# Hay un error el el tiempo de 0.02s aproximadamente, El tiempo que tarda es conocido


    def medir_n_muestras_equidistantes(self, n_muestras=10, intervalo_s=0.1, canal='A'):
        """
        Realiza una adquisición de 'n_muestras' equidistantes en el tiempo usando el CNT-91,
        devolviendo para cada muestra la frecuencia y el timestamp asociado, pudiendo elegir el canal de entrada.
    
        Parámetros:
            n_muestras: int
                Número de muestras a medir (por defecto 10)
            intervalo_s: float
                Intervalo de tiempo entre muestras en segundos (por defecto 0.1s)
            canal: str o int
                Canal de medida: 'A', 'B', 1 o 2 (por defecto 'A')
    
        Devuelve:
            lista de tuplas (frecuencia, delta_t) en floats
            (El tiempo de cada muestra es relativo al primero, es decir, delta_t = t - t0)
        """
    
        import time
    
        # ========== SECCIÓN 1: Validación y selección de canal ==========
        # Diccionario para convertir la selección del usuario al formato SCPI adecuado
        canales = {'A': '@1', 'B': '@2', '1': '@1', '2': '@2'}
        ch = str(canal).upper()
        if ch not in canales:
            raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
        canal_cmd = canales[ch]
    
        # ========== SECCIÓN 2: Configuración mínima del instrumento ==========
        # 1. Resetear el instrumento para asegurar estado limpio
        self.dev.write('*RST')
        # 2. Limpiar errores previos para evitar problemas durante la medición
        self.dev.write("*CLS")
        # 3. Configurar función de frecuencia en el canal seleccionado
        self.dev.write(f'CONF:FREQ {canal_cmd}')
        # 4. Configurar el intervalo entre muestras (apertura/pacing)
        self.dev.write(f'SENS:ACQ:APER {intervalo_s}')
        # 5. Establecer el número de muestras a adquirir (adquisición en bloque)
        self.dev.write(f'ARM:COUN {n_muestras}')
        # 6. Activar la inclusión del timestamp en la respuesta
        self.dev.write('FORM:TINF ON')
    
        # ========== SECCIÓN 3: Lanzamiento de adquisición ==========
        # 7. Iniciar la adquisición
        self.dev.write('INIT')
    
        # ========== SECCIÓN 4: Espera para completar la adquisición ==========
        # 8. Calcular y esperar el tiempo suficiente para que el instrumento termine
        #    Se añade un 10% extra como margen de seguridad
        tiempo_espera = intervalo_s * n_muestras * 1.1
        time.sleep(tiempo_espera)
    
        # ========== SECCIÓN 5: Recuperación y procesamiento de los datos ==========
        # 9. Solicitar todas las muestras disponibles en el buffer
        self.dev.write('FETC:ARR? MAX')
        data = self.dev.read()
        
    
        # 10. Procesar los datos: vienen en el formato <freq1>,<ts1>,<freq2>,<ts2>,...
        try:
            valores = [float(val) for val in data.strip().split(',') if val]
            resultados = []
            # 11. Agrupar de dos en dos: (frecuencia, timestamp)
            for i in range(0, len(valores)-1, 2):
                frecuencia = valores[i]
                timestamp = valores[i+1]
                resultados.append((frecuencia, timestamp))
            # 12. Ajustar los timestamps para que sean relativos al primero (delta_t)
            if resultados:
                t0 = resultados[0][1]
                resultados = [(f, t - t0) for (f, t) in resultados]
        except Exception:
            # Si el formato recibido no es el esperado, devolver el texto crudo para depuración
            resultados = data
    
        # 12. Devolver la lista de tuplas (frecuencia, delta_t)
        return resultados
            
 #JOAN 3 PROBAMOS OTRO METODO Usar el modo de medición con "Sample Timer"
 # 
 #   Está tiene tiempos y tiempos relativos         


    def medir_n_muestras_equidistantesV2(self, n_muestras=10, intervalo_s=0.1, canal='A'):
            """
            Versión 2.0 de la función de medición de muestras equidistantes.
            Realiza una adquisición de 'n_muestras' equidistantes en el tiempo usando el CNT-91,
            devolviendo tres arrays separados: frecuencias, timestamps y delta_tiempos.
        
            Parámetros:
                n_muestras: int
                    Número de muestras a medir (por defecto 10)
                intervalo_s: float
                    Intervalo de tiempo entre muestras en segundos (por defecto 0.1s)
                canal: str o int
                    Canal de medida: 'A', 'B', 1 o 2 (por defecto 'A')
        
            Devuelve:
                tuple: (frecuencias, timestamps, delta_tiempos)
                    - frecuencias: array de floats con las frecuencias medidas
                    - timestamps: array de floats con los tiempos absolutos
                    - delta_tiempos: array de floats con los tiempos relativos al primer valor
            """
        
            import time
            import numpy as np
        
            # ========== SECCIÓN 1: Validación y selección de canal ==========
            canales = {'A': '@1', 'B': '@2', '1': '@1', '2': '@2'}
            ch = str(canal).upper()
            if ch not in canales:
                raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
            canal_cmd = canales[ch]
        
            # ========== SECCIÓN 2: Configuración mínima del instrumento ==========
            self.dev.write('*RST')
            self.dev.write("*CLS")
            self.dev.write(f'CONF:FREQ {canal_cmd}')
            self.dev.write(f'SENS:ACQ:APER {intervalo_s}')
            self.dev.write(f'ARM:COUN {n_muestras}')
            self.dev.write('FORM:TINF ON')
        
            # ========== SECCIÓN 3: Lanzamiento de adquisición ==========
            self.dev.write('INIT')
        
            # ========== SECCIÓN 4: Espera para completar la adquisición ==========
            tiempo_espera = intervalo_s * n_muestras * 1.1
            time.sleep(tiempo_espera)
        
            # ========== SECCIÓN 5: Recuperación y procesamiento de los datos ==========
            self.dev.write('FETC:ARR? MAX')
            data = self.dev.read()
            
            try:
                # Convertir la respuesta en una lista de valores
                valores = [float(val) for val in data.strip().split(',') if val]
                
                # Separar frecuencias y timestamps
                frecuencias = valores[::2]  # Valores en posiciones pares
                timestamps = valores[1::2]  # Valores en posiciones impares
                
                # Convertir a arrays numpy
                frecuencias = np.array(frecuencias)
                timestamps = np.array(timestamps)
                
                # Calcular delta_tiempos (tiempos relativos al primer valor)
                delta_tiempos = timestamps - timestamps[0]
                
                return frecuencias, timestamps, delta_tiempos
                
            except Exception as e:
                print(f"Error procesando los datos: {str(e)}")
                return None, None, None

#JOAN 3 PROBAMOS OTRO METODO Usar el modo de medición con "Sample Timer"
# 
#   Está tiene tiempos y tiempos relativos y GRAFICAR FREQUENCIA VS TIEMPO 


    def medir_n_muestras_equidistantesV3(self, n_muestras=10, intervalo_s=0.1, canal='A', graficarFT=True):
        """
        Versión 3.0: Igual que V2, pero permite graficar resultados de frecuencia vs tiempo.
    
        Parámetros:
            n_muestras: int
                Número de muestras a medir (por defecto 10)
            intervalo_s: float
                Intervalo de tiempo entre muestras en segundos (por defecto 0.1s)
            canal: str o int
                Canal de medida: 'A', 'B', 1 o 2 (por defecto 'A')
            graficar: bool
                Si True, muestra gráfica frecuencia vs tiempo (por defecto False)
    
        Devuelve:
            tuple: (frecuencias, timestamps, delta_tiempos)
                - frecuencias: array de floats con las frecuencias medidas
                - timestamps: array de floats con los tiempos absolutos
                - delta_tiempos: array de floats con los tiempos relativos al primer valor
        """
    
        import time
        import numpy as np
    
        # ========== SECCIÓN 1: Validación y selección de canal ==========
        canales = {'A': '@1', 'B': '@2', '1': '@1', '2': '@2'}
        ch = str(canal).upper()
        if ch not in canales:
            raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
        canal_cmd = canales[ch]
    
        # ========== SECCIÓN 2: Configuración mínima del instrumento ==========
        self.dev.write('*RST')
        self.dev.write("*CLS")
        self.dev.write(f'CONF:FREQ {canal_cmd}')
        self.dev.write(f'SENS:ACQ:APER {intervalo_s}')
        self.dev.write(f'ARM:COUN {n_muestras}')
        self.dev.write('FORM:TINF ON')
    
        # ========== SECCIÓN 3: Lanzamiento de adquisición ==========
        self.dev.write('INIT')
    
        # ========== SECCIÓN 4: Espera para completar la adquisición ==========
        # Modelo combinado + margen del 10%, con casos especiales y tiempo mínimo de 0.2s
        eps = 1e-12
        T = intervalo_s
        N = n_muestras
        if abs(T - 4e-5) < eps and N == 2400:
            tiempo_espera = 0.25
        elif abs(T - 4e-4) < eps and N == 1000:
            tiempo_espera = 0.8
        else:
            raw = 2 * T**0.88 * N**0.85
            lin = 1.14 * T * N
            val = max(raw, lin)
            tiempo_espera = max(0.2, val) * 1.1
        time.sleep(tiempo_espera)
    
        # ========== SECCIÓN 5: Recuperación y procesamiento de los datos ==========
        self.dev.write('FETC:ARR? MAX')
        data = self.dev.read()
    
        try:
            # Convertir la respuesta en una lista de valores
            valores = [float(val) for val in data.strip().split(',') if val]
    
            # Separar frecuencias y timestamps
            frecuencias = np.array(valores[::2])  # Valores en posiciones pares
            timestamps = np.array(valores[1::2])  # Valores en posiciones impares
    
            # Calcular delta_tiempos (tiempos relativos al primer valor)
            delta_tiempos = timestamps - timestamps[0]
    
            # ========== SECCIÓN 6: Visualización de resultados ==========
            if graficarFT:
                import matplotlib.pyplot as plt
                from matplotlib.ticker import MaxNLocator
    
                # Estadísticas básicas
                maximo = np.max(frecuencias)
                minimo = np.min(frecuencias)
                media = np.mean(frecuencias)
                mediana = np.median(frecuencias)
                n_puntos = len(frecuencias)
    
                plt.figure(figsize=(9, 5))
                plt.plot(delta_tiempos, frecuencias, marker='o', linestyle='-', label='Frecuencia')
                plt.xlabel('Tiempo [s]', fontsize=12)
                plt.ylabel('Frecuencia [Hz]', fontsize=12)
                plt.title('Frecuencia vs Tiempo')
                plt.grid(True, which='both', linestyle='--', alpha=0.5)
                plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))
    
                # Mostrar estadísticas en la gráfica
                texto_stats = (f"Máx: {maximo:.3f} Hz\n"
                               f"Mín: {minimo:.3f} Hz\n"
                               f"Media: {media:.3f} Hz\n"
                               f"Mediana: {mediana:.3f} Hz\n"
                               f"Nº puntos: {n_puntos}")
                plt.gca().text(0.98, 0.02, texto_stats, fontsize=10,
                               ha='right', va='bottom', transform=plt.gca().transAxes,
                               bbox=dict(facecolor='white', alpha=0.7, edgecolor='gray'))
                plt.tight_layout()
                plt.show()
    
            return frecuencias, timestamps, delta_tiempos
    
        except Exception as e:
            print(f"Error procesando los datos: {str(e)}")
            return None, None, None   




#JOAN 3 PROBAMOS OTRO METODO Usar el modo de medición con "Sample Timer"
# 
#   Está tiene tiempos y tiempos relativos y GRAFICAR FREQUENCIA VS TIEMPO
# Añadido que devuelva los Allan deviation

    def medir_n_muestras_equidistantesV4(self, n_muestras=10, intervalo_s=0.1, canal='A', graficarFT=True):
        """
        Versión 4.0: Igual que V3, pero añade cálculo de Allan Deviation para diferentes Taus.
    
        Parámetros:
            n_muestras: int
                Número de muestras a medir (por defecto 10)
            intervalo_s: float
                Intervalo de tiempo entre muestras en segundos (por defecto 0.1s)
            canal: str o int
                Canal de medida: 'A', 'B', 1 o 2 (por defecto 'A')
            graficar: bool
                Si True, muestra gráfica frecuencia vs tiempo (por defecto False)
    
        Devuelve:
            tuple: (frecuencias, timestamps, delta_tiempos, allan_deviations, taus)
                - frecuencias: array de floats con las frecuencias medidas
                - timestamps: array de floats con los tiempos absolutos
                - delta_tiempos: array de floats con los tiempos relativos al primer valor
                - allan_deviations: array de floats con los valores de Allan deviation calculados
                - taus: array de floats con los valores de tau asociados a cada Allan deviation
        """
    
        import time
        import numpy as np
    
        # ========== SECCIÓN 1: Validación y selección de canal ==========
        canales = {'A': '@1', 'B': '@2', '1': '@1', '2': '@2'}
        ch = str(canal).upper()
        if ch not in canales:
            raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
        canal_cmd = canales[ch]
    
        # ========== SECCIÓN 2: Configuración mínima del instrumento ==========
        self.dev.write('*RST')
        self.dev.write("*CLS")
        self.dev.write(f'CONF:FREQ {canal_cmd}')
        self.dev.write(f'SENS:ACQ:APER {intervalo_s}')
        self.dev.write(f'ARM:COUN {n_muestras}')
        self.dev.write('FORM:TINF ON')
    
        # ========== SECCIÓN 3: Lanzamiento de adquisición ==========
        self.dev.write('INIT')
    
        # ========== SECCIÓN 4: Espera para completar la adquisición ==========
        tiempo_espera = intervalo_s * n_muestras * 1.1
        time.sleep(tiempo_espera)
    
        # ========== SECCIÓN 5: Recuperación y procesamiento de los datos ==========
        self.dev.write('FETC:ARR? MAX')
        data = self.dev.read()
    
        try:
            # Convertir la respuesta en una lista de valores
            valores = [float(val) for val in data.strip().split(',') if val]
    
            # Separar frecuencias y timestamps
            frecuencias = np.array(valores[::2])  # Valores en posiciones pares
            timestamps = np.array(valores[1::2])  # Valores en posiciones impares
    
            # Calcular delta_tiempos (tiempos relativos al primer valor)
            delta_tiempos = timestamps - timestamps[0]
    
            # ========== SECCIÓN 5B: Cálculo de Allan Deviation para diferentes Taus ==========
            N = len(frecuencias)
            allan_deviations = []
            taus = []
    
            for m in range(1, N // 2 + 1):
                M = N // m
                if M < 2:
                    break
                # Promedios de frecuencia para cada bloque de tamaño m
                promedios = [np.mean(frecuencias[i * m:(i + 1) * m]) for i in range(M)]
                # Diferencias cuadráticas entre bloques consecutivos
                dif_cuadrado = [(promedios[i + 1] - promedios[i]) ** 2 for i in range(M - 1)]
                sigma2 = np.sum(dif_cuadrado) / (2 * (M - 1))
                sigma = np.sqrt(sigma2)
                allan_deviations.append(sigma)
                taus.append(m * intervalo_s)
    
            allan_deviations = np.array(allan_deviations)
            taus = np.array(taus)
    
            # ========== SECCIÓN 6: Visualización de resultados ==========
            if graficarFT:
                import matplotlib.pyplot as plt
                from matplotlib.ticker import MaxNLocator
    
                # Estadísticas básicas
                maximo = np.max(frecuencias)
                minimo = np.min(frecuencias)
                media = np.mean(frecuencias)
                mediana = np.median(frecuencias)
                n_puntos = len(frecuencias)
    
                plt.figure(figsize=(9, 5))
                plt.plot(delta_tiempos, frecuencias, marker='o', linestyle='-', label='Frecuencia')
                plt.xlabel('Tiempo [s]', fontsize=12)
                plt.ylabel('Frecuencia [Hz]', fontsize=12)
                plt.title('Frecuencia vs Tiempo')
                plt.grid(True, which='both', linestyle='--', alpha=0.5)
                plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))
    
                # Mostrar estadísticas en la gráfica
                texto_stats = (f"Máx: {maximo:.3f} Hz\n"
                               f"Mín: {minimo:.3f} Hz\n"
                               f"Media: {media:.3f} Hz\n"
                               f"Mediana: {mediana:.3f} Hz\n"
                               f"Nº puntos: {n_puntos}")
                plt.gca().text(0.02, 0.98, texto_stats, fontsize=9,
                               ha='left', va='top', transform=plt.gca().transAxes,
                               bbox=dict(facecolor='white', alpha=0.4, edgecolor='none'))
                plt.tight_layout()
                plt.show()
    
            return frecuencias, timestamps, delta_tiempos, allan_deviations, taus
    
        except Exception as e:
            print(f"Error procesando los datos: {str(e)}")
            return None, None, None, None, None
        
        



    def medir_n_muestras_equidistantesV31(self, canal='A', n_muestras=100, intervalo_s=0.2, 
            acoplamiento=None,
            impedancia='Min',
            atenuacion=None,
            trigger_level=None,
            trigger_slope=None,
            filtro_Digital_PASSAbaja=None,
            filtro_Analog_PASSAbaja=None, 
            graficarFT=False, exportar_excel=False):
        """
        Versión 3.1: Igual que V3, pero añade opción de guardar datos en Excel (.xlsx).
    
        Parámetros:
            n_muestras: int
                Número de muestras a medir (por defecto 100)
            intervalo_s: float
                Intervalo de tiempo entre muestras en segundos (por defecto 0.2s) EL VALOR MÍNIMO ES 
            canal: str o int
                Canal de medida: 'A', 'B', 1 o 2 (por defecto 'A')
            graficarFT: bool
                Si True, muestra gráfica frecuencia vs tiempo (por defecto True)
            exportar_excel: bool
                Si True, exporta los datos a un archivo Excel .xlsx (por defecto True)
    
        Devuelve:
            tuple: (frecuencias, timestamps, delta_tiempos)
                - frecuencias: array de floats con las frecuencias medidas
                - timestamps: array de floats con los tiempos absolutos
                - delta_tiempos: array de floats con los tiempos relativos al primer valor
        """
    
        import time
        import numpy as np
    
        # ========== SECCIÓN 1: Validación y selección de canal ==========
        canales = {'A': '@1', 'B': '@2', '1': '@1', '2': '@2'}
        ch = str(canal).upper()
        if ch not in canales:
            raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
        canal_cmd = canales[ch]
    
        # ========== SECCIÓN 2: Configuración mínima del instrumento ==========
        self.dev.write('*RST')
        self.dev.write("*CLS")
        #self.dev.write('CAL:INT:AUTO ON')          # Desactiva autocalibración de interpoladores para máxima velocidad
        self.dev.write('DISP:ENAB ON')             # Apaga display para máxima velocidad
        
  
            # En cada bloque, solo se ejecuta la instrucción si el valor correspondiente NO es None.
            # ========== SECCIÓN 4: Configuración del CNT_91 ==========
            # ACOPLAMIENTO, IMPEDANCIA, ATENUACIÓN, TRIGGER LEVEL, TRIGGER SLOPE
        
        if canal in ['A', '1', '@1']:
            canal_12 = 1
        else:
            canal_12 = 2
        
        if acoplamiento is not None:
            # Default: AC, valores posibles: 'AC' o 'DC'
            self.dev.write(f':INP{canal_12}:COUP {acoplamiento}')
        if impedancia is not None:
            # Default: 1MΩ, posibles: 'MAX' (50Ω) o 'MIN' (1MΩ)
            self.dev.write(f':INP{canal_12}:IMP {impedancia}')
        if atenuacion is not None:
            # Default: 1, valores: 1–10 o 'MAX'/'MIN'
            self.dev.write(f':INP{canal_12}:ATT {atenuacion}')
            if atenuacion == 'MAX' and trigger_level is not None:
                trigger_level = float(trigger_level) * 10
        if trigger_level is not None:
            # Default: Auto (70%), valores: 0.1–10 V o 'MAX'/'MIN'
            self.dev.write(f':INP{canal_12}:LEV:AUTO OFF')
            self.dev.write(f':INP{canal_12}:LEV {trigger_level}')
        if trigger_slope is not None:
            # Default: POS, valores: 'POS' o 'NEG'
            self.dev.write(f':INP{canal_12}:SLOPe {trigger_slope}')
        if filtro_Digital_PASSAbaja is not None:
            # Filtro digital pasa bajas 1 Hz–50 MHz
            self.dev.write(f':INP{canal_12}:FILT:DIG ON')
            self.dev.write(f':INP{canal_12}:FILT:DIG:FREQ {filtro_Digital_PASSAbaja}')
        if filtro_Analog_PASSAbaja is not None:
            # Filtro analógico pasa bajas default 100 kHz
            self.dev.write(f':INP{canal_12}:FILT ON')
                
                
        self.dev.write(f"MEASure:FREQ? ({canal_cmd})")
        self.dev.write(f'SENS:ACQ:APER {intervalo_s}') # 0.004s mín
        self.dev.write(f'ARM:COUN {n_muestras}')
        
        
        
        
        
        self.dev.write('FORM:TINF ON')
        
    
        # ========== SECCIÓN 3: Lanzamiento de adquisición ==========
        self.dev.write('INIT')
    
     
        
        # ========== SECCIÓN 4: Espera para completar la adquisición ==========
        # Modelo combinado + margen del 10%, con casos especiales y tiempo mínimo de 0.2s
        eps = 1e-12
        T = intervalo_s
        N = n_muestras
        if abs(T - 4e-5) < eps and N == 2400:
            tiempo_espera = 0.25
        elif abs(T - 4e-4) < eps and N == 1000:
            tiempo_espera = 0.8
        else:
            raw = 2 * T**0.88 * N**0.85
            lin = 1.14 * T * N
            val = max(raw, lin)
            tiempo_espera = max(0.2, val) * 1.1
        time.sleep(tiempo_espera)
    
        # ========== SECCIÓN 5: Recuperación y procesamiento de los datos ==========
        self.dev.write('FETC:ARR? MAX')
        data = self.dev.read()
    
        try:
            # Convertir la respuesta en una lista de valores
            valores = [float(val) for val in data.strip().split(',') if val]
    
            # Separar frecuencias y timestamps
            frecuencias = np.array(valores[::2])  # Valores en posiciones pares
            timestamps = np.array(valores[1::2])  # Valores en posiciones impares
    
            # Calcular delta_tiempos (tiempos relativos al primer valor)
            delta_tiempos = timestamps - timestamps[0]
    
            # ========== SECCIÓN 6: Visualización de resultados ==========
            if graficarFT:
                import matplotlib.pyplot as plt
                from matplotlib.ticker import MaxNLocator
    
                maximo = np.max(frecuencias)
                minimo = np.min(frecuencias)
                media = np.mean(frecuencias)
                mediana = np.median(frecuencias)
                n_puntos = len(frecuencias)
    
                plt.figure(figsize=(9, 5))
                plt.plot(delta_tiempos, frecuencias, marker='o', linestyle='-', label='Frecuencia')
                plt.xlabel('Tiempo [s]', fontsize=12)
                plt.ylabel('Frecuencia [Hz]', fontsize=12)
                plt.title('Frecuencia vs Tiempo')
                plt.grid(True, which='both', linestyle='--', alpha=0.5)
                plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))
    
                texto_stats = (f"Máx: {maximo:.3f} Hz\n"
                               f"Mín: {minimo:.3f} Hz\n"
                               f"Media: {media:.3f} Hz\n"
                               f"Mediana: {mediana:.3f} Hz\n"
                               f"Nº puntos: {n_puntos}")
                plt.gca().text(0.98, 0.02, texto_stats, fontsize=10,
                               ha='right', va='bottom', transform=plt.gca().transAxes,
                               bbox=dict(facecolor='white', alpha=0.7, edgecolor='gray'))
                plt.tight_layout()
                plt.show()
            
            # ========== SECCIÓN NUEVA: Guardar en Excel (.xlsx) ==========
            if exportar_excel:
                import pandas as pd
                from datetime import datetime
    
                fecha_hora = datetime.now().strftime("%S_%M_%H_%d_%m_%Y")
                raw_data = {
                    "Muestra": [f"Muestra{i}" for i in range(len(frecuencias))],
                    "Frecuencia [Hz]": np.round(frecuencias, 6),   # Numérico, 6 decimales
                    "Timestamp [s]": np.round(timestamps, 2),
                    "Delta_tiempo [s]": np.round(delta_tiempos, 2)
                }
                df_raw = pd.DataFrame(raw_data)
                nombre_raw = f"RawDataFreqYTiempo_{fecha_hora}.xlsx"
                df_raw.to_excel(nombre_raw, index=False, float_format="%.6f")  # <- Guardado en xlsx
                print(f"Archivo de datos crudos guardado como: {nombre_raw}")
            self.dev.write('DISP:ENAB ON')  # Reactiva display al acabar             # Apaga display para máxima velocidad
            return frecuencias, timestamps, delta_tiempos
    
        except Exception as e:
            print(f"Error procesando los datos: {str(e)}")
            return None, None, None

    


# Version v6 pero con capacidad para configurar la medida

    def medir_n_muestras_equidistantesV7(
            self,
            n_muestras=100,
            intervalo_s=0.2,  # 4ms = 0.004s de mín recomendable ,  20 ns a 1000 s.
            canal='A',
            graficarFT=True,
            graficarDevTau=True,
            exportar_excel=True,
            configurar=False,
            impedancia=None,      # '50' (ohmios) o '1M' (megaohm)
            acoplamiento=None,    # 'DC', 'AC', 'HF', 'LF'
            atenuacion=None,      # '0' (0dB, por defecto) o '10' (10dB típico para señales grandes) 
            filtro=None,          # 'ON', 'OFF'
            triger_level=None,    # valor en voltios, e.g., 0.5
            triger_slope=None     # 'POS' (subida), 'NEG' (bajada)
        ):
        """
        Versión clásica y robusta con espera por time.sleep().
        Permite configurar impedancia, acoplamiento, atenuación, filtro, trigger level y trigger slope.
        """
        import time
        import numpy as np
    
        # ========== SECCIÓN 1: Validación y selección de canal ==========
        canales = {'A': '@1', 'B': '@2', '1': '@1', '2': '@2'}
        ch = str(canal).upper()
        if ch not in canales:
            raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
        canal_cmd = canales[ch]
        canal_num = '1' if ch in ['A', '1'] else '2'   # para los comandos INP1, INP2, etc.
    
        # ========== SECCIÓN 2: Configuración del instrumento ==========
        self.dev.write('*RST')
        self.dev.write("*CLS")
    
        # ======= SECCIÓN 2.1: Configuración extra por usuario =======
        if configurar:
            if impedancia in ['50', '1M']:
                self.dev.write(f'INP{canal_num}:IMP {impedancia}')  # Ej: INP1:IMP 50
            if acoplamiento in ['DC', 'AC', 'HF', 'LF']:
                self.dev.write(f'INP{canal_num}:COUP {acoplamiento}')  # Ej: INP1:COUP AC
            if atenuacion in ['0', '10']:
                self.dev.write(f'INP{canal_num}:ATT {atenuacion}')  # Ej: INP1:ATT 10
            if filtro in ['ON', 'OFF']:
                self.dev.write(f'INP{canal_num}:FILT {filtro}')     # Ej: INP1:FILT ON
            if triger_level is not None:
                self.dev.write(f'TRIG{canal_num}:LEV {triger_level}')  # Ej: TRIG1:LEV 0.5
            if triger_slope in ['POS', 'NEG']:
                self.dev.write(f'TRIG{canal_num}:SLOP {triger_slope}') # Ej: TRIG1:SLOP POS
    
        self.dev.write('CAL:INT:AUTO OFF')
        self.dev.write('DISP:ENAB OFF')
        self.dev.write(f'CONF:FREQ {canal_cmd}')
        self.dev.write(f'SENS:ACQ:APER {intervalo_s}')
        self.dev.write(f'ARM:COUN {n_muestras}')
        self.dev.write('FORM:TINF ON')
    
        # ========== SECCIÓN 3: Lanzamiento de adquisición y espera clásica ==========
        self.dev.write('INIT')
        # Modelo combinado + margen del 10%, con casos especiales y tiempo mínimo de 0.2s
        eps = 1e-12
        T = intervalo_s
        N = n_muestras
        if abs(T - 4e-5) < eps and N == 2400:
            tiempo_espera = 0.25
        elif abs(T - 4e-4) < eps and N == 1000:
            tiempo_espera = 0.8
        else:
            raw = 2 * T**0.88 * N**0.85
            lin = 1.14 * T * N
            val = max(raw, lin)
            tiempo_espera = max(0.2, val) * 1.1
        time.sleep(tiempo_espera)
    
        # ========== SECCIÓN 4: Recuperación y procesamiento de los datos ==========
        self.dev.write('FETC:ARR? MAX')
        data = self.dev.read()
        valores = [float(val) for val in data.strip().split(',') if val]
        if len(valores) < 2 * n_muestras:
            print(f"¡Advertencia! Recibidas menos muestras ({len(valores)//2}) de las solicitadas ({n_muestras}).")
    
        try:
            frecuencias = np.array(valores[::2])
            timestamps = np.array(valores[1::2])
            delta_tiempos = timestamps - timestamps[0]
    
            # ========== SECCIÓN 5: Cálculo de Allan Deviation ==========
            N = len(frecuencias)
            allan_deviations = []
            taus = []
            for m in range(1, N // 2 + 1):
                M = N // m
                if M < 2:
                    break
                promedios = [np.mean(frecuencias[i * m:(i + 1) * m]) for i in range(M)]
                dif_cuadrado = [(promedios[i + 1] - promedios[i]) ** 2 for i in range(M - 1)]
                sigma2 = np.sum(dif_cuadrado) / (2 * (M - 1))
                sigma = np.sqrt(sigma2)
                allan_deviations.append(sigma)
                taus.append(m * intervalo_s)
            allan_deviations = np.array(allan_deviations)
            taus = np.array(taus)
    
            # ========== SECCIÓN 6: Exportar a Excel (.xlsx, dos hojas) ==========
            if exportar_excel:
                import pandas as pd
                from datetime import datetime
    
                now = datetime.now()
                nombre_excel = (
                    f"AllanDeviation_vs_Tau_and_Freq_vs_timestamp___"
                    f"{now:%S}sec_{now:%M}min_{now:%H}hour_{now:%Y}year.xlsx"
                )
    
                raw_data = {
                    "Muestra": [f"Muestra{i}" for i in range(len(frecuencias))],
                    "Frecuencia [Hz]": np.round(frecuencias, 6),
                    "Timestamp [s]": np.round(timestamps, 6),
                    "Delta_tiempo [s]": np.round(delta_tiempos, 6)
                }
                df_raw = pd.DataFrame(raw_data)
                allan_data = {
                    "DATO": [f"DATO{i}" for i in range(len(allan_deviations))],
                    "AllanDeviation [Hz]": np.round(allan_deviations, 6),
                    "Tau [s]": np.round(taus, 6)
                }
                df_allan = pd.DataFrame(allan_data)
                with pd.ExcelWriter(nombre_excel) as writer:
                    df_raw.to_excel(writer, sheet_name='Datos Frecuencia', index=False, float_format="%.6f")
                    df_allan.to_excel(writer, sheet_name='Allan Deviation', index=False, float_format="%.6f")
                print(f"Archivo de datos guardado como: {nombre_excel}")
    
            # ========== SECCIÓN 7: Visualización de resultados (Frecuencia vs Tiempo) ==========
            if graficarFT:
                import matplotlib.pyplot as plt
                from matplotlib.ticker import MaxNLocator
    
                plt.figure(figsize=(10, 5))
                n_puntos = len(frecuencias)
                plt.scatter(delta_tiempos, frecuencias, s=6, alpha=0.7, label='Frecuencia')
                plt.xlabel('Tiempo [s]', fontsize=13)
                plt.ylabel('Frecuencia [Hz]', fontsize=13)
                plt.title('Frecuencia vs Tiempo', fontsize=15)
                plt.grid(True, which='both', linestyle='--', alpha=0.5)
                plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))
                maximo = np.max(frecuencias)
                minimo = np.min(frecuencias)
                media = np.mean(frecuencias)
                mediana = np.median(frecuencias)
                texto_stats = (f"Máx: {maximo:.2f} Hz\n"
                               f"Mín: {minimo:.2f} Hz\n"
                               f"Media: {media:.2f} Hz\n"
                               f"Mediana: {mediana:.2f} Hz\n"
                               f"Nº puntos: {n_puntos}")
                plt.gca().text(0.02, 0.98, texto_stats, fontsize=10,
                               ha='left', va='top', transform=plt.gca().transAxes,
                               bbox=dict(facecolor='white', alpha=0.3, edgecolor='none'))
                plt.tight_layout()
                plt.show()
    
            # ========== SECCIÓN 8: Visualización de resultados (Allan Deviation vs Tau) ==========
            if graficarDevTau:
                import matplotlib.pyplot as plt
    
                plt.figure(figsize=(10, 5))
                plt.scatter(taus, allan_deviations, s=18, color='C0', alpha=0.8, label='Adev')
                plt.xscale('log')
                plt.yscale('log')
                plt.xlabel(r'$\tau$ [s]', fontsize=13)
                plt.ylabel('Allan Deviation [Hz]', fontsize=13)
                plt.title('Allan Deviation vs Tau', fontsize=15)
                plt.grid(True, which='both', linestyle='--', alpha=0.45)
    
                idx_min = np.argmin(allan_deviations)
                tau_min = taus[idx_min]
                adev_min = allan_deviations[idx_min]
                plt.scatter([tau_min], [adev_min], color='red', s=70, label=f'Mín Adev\nTau={tau_min:.2f}s\nAdev={adev_min:.2f}Hz', zorder=5)
                plt.legend(fontsize=10)
                plt.annotate(f'Mín:\nTau={tau_min:.2f}s\nAdev={adev_min:.2f}Hz',
                             xy=(tau_min, adev_min), xytext=(0.05, 0.98),
                             textcoords='axes fraction', ha='left', va='top',
                             fontsize=10, color='red',
                             bbox=dict(facecolor='white', alpha=0.45, edgecolor='red'))
                plt.tight_layout()
                plt.show()
    
            self.dev.write('DISP:ENAB ON')  # Reactiva display al acabar
    
            return frecuencias, timestamps, delta_tiempos, allan_deviations, taus
    
        except Exception as e:
            print(f"Error procesando los datos: {str(e)}")
            return None, None, None, None, None
    



# Generá mas error que block measurement


    def continuous_measurament_v31(self, n_muestras=100, intervalo_s=0.2, canal='A', graficarFT=True, exportar_excel=True):
        """
        Medición continua: inicia la medición en modo continuo, espera el tiempo necesario,
        hace un ABORT, luego recupera exactamente n_muestras.
        Exporta a Excel (.xlsx) y puede graficar si se desea.
        """
        import time
        import numpy as np
    
        # ========== SECCIÓN 1: Validación y selección de canal ==========
        canales = {'A': '@1', 'B': '@2', '1': '@1', '2': '@2'}
        ch = str(canal).upper()
        if ch not in canales:
            raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
        canal_cmd = canales[ch]
    
        # ========== SECCIÓN 2: Configuración mínima del instrumento ==========
        self.dev.write('*RST')
        self.dev.write("*CLS")
        self.dev.write('CAL:INT:AUTO OFF')
        self.dev.write('DISP:ENAB OFF')
        self.dev.write(f'CONF:FREQ {canal_cmd}')
        self.dev.write(f'SENS:ACQ:APER {intervalo_s}')  # Tiempo de apertura por muestra
        self.dev.write("ARM:COUNT INF")
        self.dev.write('ARM:CONT ON')  # ARM Continuous mode ON (medición continua)
        self.dev.write('FORM:TINF ON')  # Formato con timestamps
    
        # ========== SECCIÓN 3: Lanzamiento de medición continua ==========
        self.dev.write('INIT')
        # Espera suficiente para recoger todas las muestras + margen
        # Modelo combinado + margen del 10%, con casos especiales y tiempo mínimo de 0.2s
        eps = 1e-12
        T = intervalo_s
        N = n_muestras
        if abs(T - 4e-5) < eps and N == 2400:
            tiempo_espera = 0.25
        elif abs(T - 4e-4) < eps and N == 1000:
            tiempo_espera = 0.8
        else:
            raw = 2 * T**0.88 * N**0.85
            lin = 1.14 * T * N
            val = max(raw, lin)
            tiempo_espera = max(0.2, val) * 1.1
        time.sleep(tiempo_espera)
        
        
        
        
        
    
        # ========== SECCIÓN 4: Abortamos y recuperamos muestras ==========
        self.dev.write('ABOR')  # Aborta la medición continua
        self.dev.write(f'FETC:ARR? {n_muestras}')  # Recupera sólo n_muestras
    
        data = self.dev.read()
    
        try:
            # Convertir la respuesta en una lista de valores
            valores = [float(val) for val in data.strip().split(',') if val]
    
            if len(valores) < 2 * n_muestras:
                print(f"¡Advertencia! Recibidas menos muestras ({len(valores)//2}) de las solicitadas ({n_muestras}).")
    
            # Separar frecuencias y timestamps
            frecuencias = np.array(valores[::2])  # Valores en posiciones pares
            timestamps = np.array(valores[1::2])  # Valores en posiciones impares
    
            # Calcular delta_tiempos (tiempos relativos al primer valor)
            delta_tiempos = timestamps - timestamps[0]
    
            # ========== SECCIÓN 5: Visualización de resultados ==========
            if graficarFT:
                import matplotlib.pyplot as plt
                from matplotlib.ticker import MaxNLocator
    
                maximo = np.max(frecuencias)
                minimo = np.min(frecuencias)
                media = np.mean(frecuencias)
                mediana = np.median(frecuencias)
                n_puntos = len(frecuencias)
    
                plt.figure(figsize=(9, 5))
                plt.plot(delta_tiempos, frecuencias, marker='o', linestyle='-', label='Frecuencia')
                plt.xlabel('Tiempo [s]', fontsize=12)
                plt.ylabel('Frecuencia [Hz]', fontsize=12)
                plt.title('Frecuencia vs Tiempo')
                plt.grid(True, which='both', linestyle='--', alpha=0.5)
                plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))
    
                texto_stats = (f"Máx: {maximo:.3f} Hz\n"
                               f"Mín: {minimo:.3f} Hz\n"
                               f"Media: {media:.3f} Hz\n"
                               f"Mediana: {mediana:.3f} Hz\n"
                               f"Nº puntos: {n_puntos}")
                plt.gca().text(0.98, 0.02, texto_stats, fontsize=10,
                               ha='right', va='bottom', transform=plt.gca().transAxes,
                               bbox=dict(facecolor='white', alpha=0.7, edgecolor='gray'))
                plt.tight_layout()
                plt.show()
    
            # ========== SECCIÓN 6: Guardar en Excel (.xlsx) ==========
            if exportar_excel:
                import pandas as pd
                from datetime import datetime
    
                fecha_hora = datetime.now().strftime("%S_%M_%H_%d_%m_%Y")
                raw_data = {
                    "Muestra": [f"Muestra{i}" for i in range(len(frecuencias))],
                    "Frecuencia [Hz]": np.round(frecuencias, 6),
                    "Timestamp [s]": np.round(timestamps, 2),
                    "Delta_tiempo [s]": np.round(delta_tiempos, 2)
                }
                df_raw = pd.DataFrame(raw_data)
                nombre_raw = f"RawDataFreqYTiempo_CONT_{fecha_hora}.xlsx"
                df_raw.to_excel(nombre_raw, index=False, float_format="%.6f")
                print(f"Archivo de datos crudos guardado como: {nombre_raw}")
    
            self.dev.write('DISP:ENAB ON')  # Reactiva display al acabar
    
            return frecuencias, timestamps, delta_tiempos
    
        except Exception as e:
            print(f"Error procesando los datos: {str(e)}")
            return None, None, None
    
 

#### Saca el ADEV y Estádisticas de una única cuenta
    def calcular_Adev_Estadistics(self, canal='A',N_muestras = 100, intervalo_captura=0.0002, pacing_time=None, acoplamiento='AC', impedancia='MIN', atenuacion=None, trigger_level=None, trigger_slope=None, filtro=None):
        
        import time
        try:
            # ========== SECCIÓN 1: Validación y selección de canal ==========
            canales = {'A': '@1', 'B': '@2', '1': '@1', '2': '@2'}
            ch = str(canal).upper()
            if ch not in canales:
                raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
            canal_cmd = canales[ch]
            
            # ========== SECCIÓN 2: Resetear y limpiar instrumento ==========
            self.dev.write('*RST')
            self.dev.write('*CLS')
            
                
            # ========== SECCIÓN 3: Configuración del CNT_91  ACOPLAMIENTO, IMPEDANCIA, ATENUACIÓN, TRIGGER LEVEL, TRIGGER SLOPE ==========
            
            # En cada bloque, solo se ejecuta la instrucción si el valor correspondiente NO es None.
            
            # Configura acoplamiento (AC/DC)
            if acoplamiento is not None:
                # Default value: AC  , Posible valores: AC or DC
                self.dev.write(f':INP{canal_cmd}:COUP {acoplamiento}')
           
            # Configura impedancia (50Ω–1MΩ, MAX, MIN)
            if impedancia is not None:
                # Default value: 50 Ohm , Posible valores: between [50 Ohm  1M Ohm] or MAX or MIN
                self.dev.write(f':INP{canal_cmd}:IMP {impedancia}')
            
            # Configura atenuación (0 dB–10 dB, MAX, MIN)
            if atenuacion is not None:
                # Default value: 0 dB , Posible valores: between [0x to 10x] or MAX or MIN , 
                # <Numeric values> ≤ 5 → atenuación 1, <Numeric values> > 5 → atenuación 10.
                self.dev.write(f':INP{canal_cmd}:ATT {atenuacion}')
            
            # Configura nivel de trigger (0.1 V–10 V, MAX, MIN)
            if trigger_level is not None:
                # Default value: 0.5 V , Posible valores: between [0.1 V to 10 V] or MAX or MIN
                self.dev.write(f':INP{canal_cmd}:TRL {trigger_level}')
            
            # Configura pendiente de trigger (POS/NEG)
            if trigger_slope is not None:
                # Default value: POS   , Posible valores: POS or NEG
                self.dev.write(f':INP{canal_cmd}:TRS {trigger_slope}')
            
            # Configura frecuencia de filtro digital (1 Hz–50 MHz, MAX, MIN)
            if filtro is not None:
                # Default value: 100e3 Hz   , Posible valores: between [1 to 50e6 Hz] or MAX or MIN
                self.dev.write(f':INP{canal_cmd}:FIL:DIG:FREQ {filtro}')     
                
        
            # ========== SECCIÓN 3: Configuración de canal y adquisición ==========
            self.dev.write(f':CONF:FREQ {canal_cmd}')
            self.dev.write(f'SENS:ACQ:APER  {intervalo_captura}')  # Tiempo de apertura por muestra default: 0.2 s , posible valores: entre [ _____________ s]
            
            # ========== SECCIÓN 4: Configuración de estadística ADEV ==========
            self.dev.write(':CALC:AVER:STAT ON')
            self.dev.write(':CALC:TYPE ADEV')
            self.dev.write(f':ARM:START:COUN N_muestras')  # Número de muestras default: 100 , posible valores: entre [1 to 1000000]
            
            if pacing_time is not None:
                self.dev.write(f'TRIGger:SOURce TIMer')
                self.dev.write(f':TRIG:TIM {pacing_time}') # Tiempo entre muestras , default: 0.2 s , posible valores: entre [____________ s]
            
            
            # ========== SECCIÓN 5: Iniciar medición ==========
            self.dev.write(':INIT')
            
            ### Esperamos a que la medición termine
            current = 0
            while float(current) != N_muestras:
                self.dev.write(':CALCulate:AVERage:COUNt:CURRent?')
                current = self.dev.read()
            
            
            # ========== SECCIÓN 6: Lectura de Allan deviation ==========
            
            self.dev.write(':CALC:DATA?')
            resp_adev = self.dev.read()
            try:
                valores = [float(val) for val in resp_adev.strip().split(',') if val]
                allan_deviation = valores[0] if len(valores) >= 1 else None
            except Exception:
                allan_deviation = None
            
            # ========== SECCIÓN 7: Lectura de estadísticas de media ==========
            self.dev.write(':CALC:AVER:ALL?')
            resp_estadisticas = self.dev.read()
            try:
                valores = [float(val.strip()) for val in resp_estadisticas.strip().split(',') if val.strip()]
                valor_medio       = valores[0] if len(valores) > 0 else None
                desviacion_tipica = valores[1] if len(valores) > 1 else None
                valor_minimo      = valores[2] if len(valores) > 2 else None
                valor_maximo      = valores[3] if len(valores) > 3 else None
            except Exception:
                valor_medio       = None
                desviacion_tipica = None
                valor_minimo      = None
                valor_maximo      = None
                
                
            
            self.dev.write(':CALC:AVER:STAT OFF') ## QUITAR ESTA INSTRUCCIÓN SI TE INTERESA QUE SIGA MIDIENDO
            
            # ========== SECCIÓN 8: Devolver resultados ==========
            return allan_deviation, valor_medio, desviacion_tipica, valor_minimo, valor_maximo
    
        except Exception as e:
            print(f"Error al calcular ADEV y media: {str(e)}")
            return None, None, None, None, None



    #### Saca el ADEVs y Estádisticas en Bloque de varias mediciones desde int min a int max

    def calcular_Adev_Estadistics_improved(
        self,
        canal='A',
        N_muestras=100,
        intervalo_captura_min=0.004E-3,
        intervalo_captura_max=10E-3,
        pasos=6,
        pacing_time=None,
        acoplamiento=None,
        impedancia=None,
        atenuacion=None,
        trigger_level=None,
        trigger_slope=None,
        filtro_Digital_PASSAbaja=None,
        filtro_Analog_PASSAbaja=None
    ):
        import time
        import numpy as np
    
        def _format_tau(tau: float) -> str:
            """
            Formatea el valor de tau de modo que:
              - Si tau < 1e-3 → 1 cifra significativa
              - Si tau >= 1e-3 → 2 cifras significativas
            """
            if tau < 1e-3:
                return f"{tau:.2g}"
            else:
                return f"{tau:.3g}"
    
        try:
            # ========== SECCIÓN 1: Validación y selección de canal ==========
            canales = {'A': '@1', 'B': '@2', '1': '@1', '2': '@2'}
            ch = str(canal).upper()
            if ch not in canales:
                raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
            canal_cmd = canales[ch]
           
            if ch in ['A', '1', '@1']:
                canal_12 = 1
            else:
                canal_12 = 2
    
            # ========== SECCIÓN 2: Resetear y limpiar instrumento ==========
            self.dev.write('*RST')
            self.dev.write('*CLS')
            
            # ========== SECCIÓN 3: Configuración del CNT_91:
            # ACOPLAMIENTO, IMPEDANCIA, ATENUACIÓN, TRIGGER LEVEL, TRIGGER SLOPE ==========
            # Configura acoplamiento (AC/DC)
            if acoplamiento is not None:
                # Default value: AC  , Posible valores: 'AC' or 'DC'
                self.dev.write(f':INP{canal_12}:COUP {acoplamiento}')
                
           
            # Configura impedancia (50Ω–1MΩ, MAX, MIN)
            if impedancia is not None:
                # Default value: 1M Ohm , Posible valores: Only string values ['MAX' = 50 Ohm  , 'MIN' = 1M Ohm] 
                self.dev.write(f':INP{canal_12}:IMP {impedancia}')
            
            # Configura atenuación (0 dB–10 dB, MAX, MIN)
            if atenuacion is not None:
                # Default value: 0 dB , Posible valores: between [0x to 10x] or MAX or MIN , 
                # <Numeric values> ≤ 5 → atenuación 1, <Numeric values> > 5 → atenuación 10.
                self.dev.write(f':INP{canal_12}:ATT {atenuacion}')
                if atenuacion == 10:
                    trigger_level = trigger_level*10
            
            # Configura nivel de trigger (0.1 V–10 V, MAX, MIN)
            if trigger_level is not None:
                # Default value: 0.5 V , Posible valores: between [0.1 V to 10 V] or MAX or MIN
                self.dev.write(f':INP{canal_12}:LEV:AUTO OFF')
                self.dev.write(f':INP{canal_12}:LEV {trigger_level}') # El trigger_level = trigger_level*atenuacion
            
            # Configura pendiente de trigger (POS/NEG)
            if trigger_slope is not None:
                # Default value: POS   , Posible valores: POS or NEG
                self.dev.write(f':INP{canal_12}:SLOPe {trigger_slope}')
            
            # Configura frecuencia de filtro digital (1 Hz–50 MHz, MAX, MIN) low Pass " PASA BAJAS "
            if filtro_Digital_PASSAbaja is not None:
                self.dev.write(f':INP{canal_12}:FILT:DIG ON') # Activación filtro digital
                # Default value: 100e3 Hz   , Posible valores: between [1 to 50e6 Hz] or MAX or MIN
                self.dev.write(f':INP{canal_12}:FILT:DIG:FREQ {filtro_Digital_PASSAbaja}') # Frequencia de corte pasa bajas

            # Configura frecuencia de filtro Analógico LowPass " PASA BAJAS " freq corte = 100Khz
            if filtro_Analog_PASSAbaja is not None:
                self.dev.write(f':INP{canal_12}:FILT ON') # Activación filtro aNALOG, se puede usar con el de arriba

            # ========== SECCIÓN 4: Configuración de canal y preparación de ADEV ==========
            
    
            self.dev.write(':CALC:AVER:STAT ON')
            self.dev.write(':CALC:AVER:TYPE ADEV')
    

    
            # ========== SECCIÓN 5: Loop de MULTIPLES INTERVALOS DE CAPTURA ==========
            resultados = []
    
            if pasos < 1:
                raise ValueError("El parámetro 'pasos' debe ser al menos 1")
            if pasos == 1:
                valores_intervalo = [intervalo_captura_min]
            else:
                valores_intervalo = np.logspace(
                    np.log10(intervalo_captura_min),
                    np.log10(intervalo_captura_max),
                    pasos
                )
    
            for intervalo_captura in valores_intervalo:
                # Formatear tau según su tamaño
                tau_str = _format_tau(intervalo_captura)
                self.dev.write(f':SENS:ACQ:APER {tau_str}') # MIN gives 20 ns and MAX gives 1000 s.

    
                self.dev.write(f':ARM:START:COUN {N_muestras}')
                if pacing_time is not None:
                    self.dev.write(f'TRIGger:SOURce TIMer')
                    self.dev.write(f':TRIG:TIM {pacing_time}') #esta en s between 2 micro and 500s 
                
                self.dev.write(':INIT')
    
                current = 0
                while float(current) != N_muestras:
                    self.dev.write(':CALCulate:AVERage:COUNt:CURRent?')
                    current = self.dev.read()
                time.sleep(0.4)
                self.dev.write(':CALC:DATA?')
                resp_adev = self.dev.read()
                time.sleep(0.4)
                try:
                    vals_adev = [float(val) for val in resp_adev.strip().split(',') if val]
                    allan_deviation = vals_adev[0] if len(vals_adev) >= 1 else None
                except Exception:
                    allan_deviation = None
    
                self.dev.write(':CALC:AVER:ALL?')
                resp_estad = self.dev.read()
                try:
                    vals_est = [
                        float(val.strip())
                        for val in resp_estad.strip().split(',')
                        if val.strip()
                    ]
                    valor_medio = vals_est[0] if len(vals_est) > 0 else None
                    desviacion_tipica = vals_est[1] if len(vals_est) > 1 else None
                    valor_minimo = vals_est[2] if len(vals_est) > 2 else None
                    valor_maximo = vals_est[3] if len(vals_est) > 3 else None
                except Exception:
                    valor_medio = None
                    desviacion_tipica = None
                    valor_minimo = None
                    valor_maximo = None
    
                resultados.append(
                    (
                        intervalo_captura,
                        allan_deviation,
                        valor_medio,
                        desviacion_tipica,
                        valor_minimo,
                        valor_maximo
                    )
                )
    
            # ========== SECCIÓN 6: Desactivar estadística ADEV ==========
            
    
            # ========== SECCIÓN 7: Devolver resultados ==========
            return resultados
    
        except Exception as e:
            print(f"Error en calcular_Adev_Estadistics_improved: {str(e)}")
            return []



    def calc_Adev_Estadistics1(
            self,
            BTB = False,
            canal='A',
            N_muestras=100,
            intervalo_captura_min=0.004E-3,
            intervalo_captura_max=10E-3,
            pasos=6,
            pacing_time=None,
            acoplamiento=None,
            impedancia=None,
            atenuacion=None,
            trigger_level=None,
            trigger_slope=None,
            filtro_Digital_PASSAbaja=None,
            filtro_Analog_PASSAbaja=None,
            guardar=False
        ):
        import time
        import numpy as np
        # ========== SECCIÓN 0: Medir frecuencia del canal ==========
        try:
            frequencia_medicion = self.measure_frequency(canal)
            try:
                frequencia_medicion = float(frequencia_medicion)
            except Exception:
                pass  # Si la conversión falla, dejamos el valor original
        except Exception as e:
            frequencia_medicion = None
            print("No se pudo medir la frecuencia:", e)
    
        # ========== SECCIÓN 1: Preparar Excel si se solicita guardar ==========
        if guardar:
            import os
            from datetime import datetime
            from openpyxl import Workbook
    
            carpeta = 'Mediciones_CNT'
            os.makedirs(carpeta, exist_ok=True)
            now = datetime.now()
            timestamp = f'D{now.day:02d}_M{now.month:02d}_Y{now.year}_h{now.hour:02d}_m{now.minute:02d}_s{now.second:02d}'
            nombre_archivo = os.path.join(carpeta, f"Mediciones_estadisticas_{timestamp}.xlsx")
    
            wb = Workbook()
            ws = wb.active
            ws.title = 'Mediciones'
            # Título y frecuencia medida (se actualizará tras medir)
            ws.append(['[  Mediciones estadisticas a una frecuencia ',frequencia_medicion, ' Hz  ]'])
            
            # Configuración de parámetros (si None, se muestra valor por defecto real del equipo)
            ws.append(['Configuración:'])
    
            # Definición de valores por defecto reales del CNT-91
            coupling_val    = acoplamiento if acoplamiento is not None else 'AC'
            impedance_val   = '50'  if impedancia   is not None else '1M'
            attenuation_val = atenuacion   if atenuacion   is not None else 1
            trigger_lvl_val = trigger_level if trigger_level is not None else 'Auto (70%)'
            trigger_slp_val = trigger_slope if trigger_slope is not None else 'POS'
            filtro_dig_val  = filtro_Digital_PASSAbaja if filtro_Digital_PASSAbaja is not None else 'OFF'
            filtro_ana_val  = filtro_Analog_PASSAbaja  if filtro_Analog_PASSAbaja  is not None else 'OFF'
            pacing_time_val  = pacing_time  if pacing_time  is not None else '20m'
            ws.append(['canal',                    str(canal)+' Ch'])
            ws.append(['N_muestras',               N_muestras])
            ws.append(['intervalo_captura_min',    str(intervalo_captura_min) +' s'])
            ws.append(['intervalo_captura_max',    str(intervalo_captura_max) +' s'])
            ws.append(['pasos',                    str(pasos) +' pasos'])
            ws.append(['pacing_time',              str(pacing_time_val) +' s'])
            ws.append(['acoplamiento',             coupling_val])
            ws.append(['impedancia',               str(impedance_val) + 'Ohm'])
            ws.append(['atenuacion',               str(attenuation_val) + 'x'])
            ws.append(['trigger_level',            str(trigger_lvl_val) + ' V'])
            ws.append(['trigger_slope',            trigger_slp_val])
            ws.append(['filtro_Digital_PASSAbaja', str(filtro_dig_val) + ' Hz'])
            ws.append(['filtro_Analog_PASSAbaja',  str(filtro_ana_val)])
            ws.append([])
            # Encabezados de la tabla de resultados
            headers = [
                'Numero de medida',                # << NUEVA COLUMNA >>
                'intervalo_captura[s]', 'allan_deviation[Hz]',
                'valor_medio[Hz]',       'desviacion_tipica[Hz]',
                'valor_minimo[Hz]',      'valor_maximo[Hz]',
                'Estabilidad_Adev[-]',  'frequencia_of_medicion[Hz]'
            ]
            ws.append(headers)
    
        # ========== SECCIÓN 2: Resetear y limpiar instrumento ==========
        self.dev.write('*RST')
        self.dev.write('*CLS')
    
        def _format_tau(tau: float) -> str:
            """
            Formatea el valor de tau de modo que:
              - Si tau < 1e-3 → 1 cifra significativa
              - Si tau >= 1e-3 → 2 cifras significativas
            """
            if tau < 1e-3:
                return f"{tau:.2g}"
            else:
                return f"{tau:.3g}"
    
        # ========== SECCIÓN 3: Validación y selección de canal ==========
        try:
            canales = {'A': '@1', 'B': '@2', '1': '@1', '2': '@2'}
            ch = str(canal).upper()
            if ch not in canales:
                raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
            canal_cmd = canales[ch]
            canal_12 = 1 if ch in ['A', '1', '@1'] else 2
    
            # ========== SECCIÓN 4: Configuración del CNT_91 ==========
            # ACOPLAMIENTO, IMPEDANCIA, ATENUACIÓN, TRIGGER LEVEL, TRIGGER SLOPE
            if acoplamiento is not None:
                # Default: AC, valores posibles: 'AC' o 'DC'
                self.dev.write(f':INP{canal_12}:COUP {acoplamiento}')
            if impedancia is not None:
                # Default: 1MΩ, posibles: 'MAX' (50Ω) o 'MIN' (1MΩ)
                self.dev.write(f':INP{canal_12}:IMP {impedancia}')
            if atenuacion is not None:
                # Default: 1, valores: 1–10 o 'MAX'/'MIN'
                self.dev.write(f':INP{canal_12}:ATT {atenuacion}')
                if atenuacion == 'MAX' and trigger_level is not None:
                    trigger_level = float(trigger_level) * 10
            if trigger_level is not None:
                # Default: Auto (70%), valores: 0.1–10 V o 'MAX'/'MIN'
                self.dev.write(f':INP{canal_12}:LEV:AUTO OFF')
                self.dev.write(f':INP{canal_12}:LEV {trigger_level}')
            if trigger_slope is not None:
                # Default: POS, valores: 'POS' o 'NEG'
                self.dev.write(f':INP{canal_12}:SLOPe {trigger_slope}')
            if filtro_Digital_PASSAbaja is not None:
                # Filtro digital pasa bajas 1 Hz–50 MHz
                self.dev.write(f':INP{canal_12}:FILT:DIG ON')
                self.dev.write(f':INP{canal_12}:FILT:DIG:FREQ {filtro_Digital_PASSAbaja}')
            if filtro_Analog_PASSAbaja is not None:
                # Filtro analógico pasa bajas default 100 kHz
                self.dev.write(f':INP{canal_12}:FILT ON')
            self.dev.write(f"MEASure:FREQ? ({canal_cmd})")
            # ========== SECCIÓN 5: Configuración de canal y preparación de ADEV ==========
            if BTB is not False:
                cmd = f":MEASure:ARRay:FREQ:BTBack? {N_muestras},({canal_cmd})"
            self.dev.write(cmd)
            self.dev.write(':CALC:AVER:STAT ON')
            self.dev.write(':CALC:AVER:TYPE ADEV')

            # ========== SECCIÓN 6: Loop de MULTIPLES INTERVALOS DE CAPTURA ==========
            resultados = []
            if pasos < 1:
                raise ValueError("El parámetro 'pasos' debe ser al menos 1")
            valores_intervalo = (
                [intervalo_captura_min] if pasos == 1 else
                np.logspace(
                    np.log10(intervalo_captura_min),
                    np.log10(intervalo_captura_max),
                    pasos
                )
            )
    
            for idx, intervalo_captura in enumerate(valores_intervalo):
                tau_str = _format_tau(intervalo_captura)
                self.dev.write(f':SENS:ACQ:APER {tau_str}')
                self.dev.write(f':ARM:START:COUN {N_muestras}')
                if pacing_time is not None:
                    self.dev.write(f'TRIGger:SOURce TIMer')
                    self.dev.write(f':TRIG:TIM {pacing_time}')
                self.dev.write(':INIT')
    
                current = 0
                while float(current) != N_muestras:
                    self.dev.write(':CALCulate:AVERage:COUNt:CURRent?')
                    current = self.dev.read()
                time.sleep(0.4)
                self.dev.write(':CALC:DATA?')
                resp_adev = self.dev.read()
                time.sleep(0.4)
                try:
                    vals_adev = [float(val) for val in resp_adev.strip().split(',') if val]
                    allan_deviation = vals_adev[0] if vals_adev else None
                except Exception:
                    allan_deviation = None
    
                self.dev.write(':CALC:AVER:ALL?')
                resp_estad = self.dev.read()
                try:
                    vals_est = [float(val.strip()) for val in resp_estad.strip().split(',') if val.strip()]
                    valor_medio, desviacion_tipica, valor_minimo, valor_maximo = (
                        vals_est + [None, None, None, None]
                    )[:4]
                except Exception:
                    valor_medio = desviacion_tipica = valor_minimo = valor_maximo = None
    
                try:
                    if allan_deviation is not None and isinstance(frequencia_medicion, (int, float)):
                        Estabilidad_Adev = allan_deviation / frequencia_medicion
                    else:
                        Estabilidad_Adev = None
                except Exception:
                    Estabilidad_Adev = None
    
                resultado_dict = {
                    'Numero de medida':       idx,
                    'intervalo_captura':     intervalo_captura,
                    'allan_deviation':       allan_deviation,
                    'valor_medio':           valor_medio,
                    'desviacion_tipica':     desviacion_tipica,
                    'valor_minimo':          valor_minimo,
                    'valor_maximo':          valor_maximo,
                    'Estabilidad_Adev':      Estabilidad_Adev,
                    'frequencia_of_medicion':frequencia_medicion
                }
                resultados.append(resultado_dict)
    
                # Añadir fila al Excel si guardar=True
                if guardar:
                    ws.append([
                        idx,
                        intervalo_captura, allan_deviation,
                        valor_medio,      desviacion_tipica,
                        valor_minimo,     valor_maximo,
                        Estabilidad_Adev, frequencia_medicion
                    ])
    
            # ========== SECCIÓN 7: Desactivar estadística ADEV ==========
            # (Puedes desactivar aquí si lo consideres necesario)
            # self.dev.write(':CALC:AVER:STAT OFF')
    
            return resultados
    
        except Exception as e:
            print(f"Error en calcular_Adev_Estadistics: {str(e)}")
            return []
    
        finally:
            if guardar and 'wb' in locals():
                wb.save(nombre_archivo)
                print(f"Archivo de mediciones guardado en: {nombre_archivo}")




# comparar con block measurament
        
    def medir_n_muestras_equidistantesBTBack(
                self,
                n_muestras=100,
                canal='A',
                intervalo_captura=0.2,
                graficarFT=False,
                exportar_excel=False
            ):
        """
        Adquisición rápida de un array de frecuencias y sus timestamps usando los comandos:
        - :MEASure:ARRay:FREQuency:BTBack? N,canal
        - :MEASure:ARRay:TSTAmp? N,canal
    
        Permite especificar el tiempo de integración (apertura) para cada medición.
    
        Parámetros de entrada:
        ----------------------
        n_muestras : int
            Número de muestras a adquirir (default: 100)
        canal : str o int
            Canal de medida: 'A', 'B', 1 o 2 (default: 'A')
        intervalo_captura : float o None
            Tiempo de integración (apertura) en segundos para cada muestra.
            - Mínimo típico: 4e-6 s (4 µs)
            - Máximo típico: 1000 s (depende del instrumento)
            - Default CNT-91: 0.2 s
            Si se pasa None, NO se configura el tiempo de integración y se usa el valor actual del instrumento.
        graficarFT : bool
            Si True, grafica frecuencia vs tiempo relativo (default: False)
        exportar_excel : bool
            Si True, exporta los datos a un archivo Excel (default: False)
    
        Salida:
        -------
        frecuencias : np.ndarray
            Array de frecuencias medidas (Hz)
        timestamps : np.ndarray
            Array de tiempos absolutos (s) en los que se tomó cada muestra
        delta_tiempos : np.ndarray
            Array de tiempos relativos al inicio (s)
    
        Notas sobre el funcionamiento y precisión:
        ------------------------------------------
        - El tiempo de integración real puede diferir ligeramente del solicitado, especialmente para valores muy pequeños (<1 ms) o muy grandes.
        - El instrumento realiza las mediciones en modo "zero dead time" (sin huecos entre muestras) **siempre que el intervalo de captura lo permita**. Para valores muy pequeños, puede haber limitaciones por el hardware.
        - Los timestamps devueltos por :MEASure:ARRay:TSTAmp? reflejan el tiempo real de cada medición, pero pueden tener un pequeño error (jitter) debido a la resolución interna del instrumento (~nanosegundos a microsegundos).
        - Si el intervalo de captura es muy pequeño, el instrumento puede no ser capaz de mantener "zero dead time" y aparecerán pequeños huecos.
        - Si el usuario pasa None como intervalo_captura, se usa el valor actual configurado en el instrumento.
    
        Errores esperados:
        ------------------
        - Si el instrumento no responde en 30 segundos, se lanza un timeout.
        - Si el número de frecuencias y timestamps no coincide, se lanza un error.
        
        
        
        
        
        Set Pacing Time
 :TRIGger: TIMer
 <Numeric value> | MIN | MAX
 This command sets the sample rate, for instance in conjunction with the statistics
 functions.
 Parameters:
 <Numeric value> is a time length between 2 s and 500 s, entered in seconds.
 MIN means 2 s.
 MAXmeans 500 s.
 Returned format: <Numeric value>
 *RST condition: 20 ms
        
        
        """
    
        import numpy as np
        import time
    
        # ====== SECCIÓN 1: Validación de parámetros y canal ======
        canales = {'A': 1, 'B': 2, '1': 1, '2': 2}
        ch = str(canal).upper()
        if ch not in canales:
            raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
        canal_num = canales[ch]
    
        # ====== SECCIÓN 2: Reset y limpieza del instrumento ======
        self.dev.write('*RST')
        self.dev.write('*CLS')
        
        self.dev.write(f":FORMat:TINFormation ON")
        
      
        # ====== SECCIÓN 4: Configuración del tiempo de integración (apertura) ======
        # Si el usuario pasa None, NO se configura y se usa el valor actual del instrumento
        if intervalo_captura is not None:
            # Valores típicos CNT-91: mínimo 4e-6 s, máximo 1000 s, default 0.2 s
            self.dev.write(f":SENS:ACQ:APER {intervalo_captura}")
            
            
        self.dev.write(f":MEASure:ARRay:TSTAmp? {n_muestras},{canal_num}")
        
        data_time = self.dev.read()
        timestamps = np.array([float(val) for val in data_time.strip().split(',') if val])
           
        # ====== SECCIÓN 5: Adquisición de frecuencias y timestamps ======
        # El instrumento realiza una sola adquisición de N muestras en modo "zero dead time" si es posible.
        # Los timestamps corresponden exactamente a las frecuencias adquiridas.
        self.dev.write(f":MEASure:ARRay:FREQuency:BTBack? {n_muestras},{canal_num}")
        
        data_freq = self.dev.read()
        frecuencias = np.array([float(val) for val in data_freq.strip().split(',') if val])
        

    
        # ====== SECCIÓN 6: Corrección de timestamps ======
        # Nos quedamos solo con los valores útiles (índices impares: 1, 3, 5, ...)
        timestamps_utiles = timestamps[1::2]
        if len(frecuencias) != len(timestamps_utiles):
            raise RuntimeError("El número de frecuencias y timestamps útiles no coincide. Revisa el formato SCPI.")
        delta_tiempos = timestamps_utiles - timestamps_utiles[0]
    
        # ====== SECCIÓN 7: Visualización opcional ======
        if graficarFT:
            import matplotlib.pyplot as plt
            from matplotlib.ticker import MaxNLocator
    
            maximo = np.max(frecuencias)
            minimo = np.min(frecuencias)
            media = np.mean(frecuencias)
            mediana = np.median(frecuencias)
            n_puntos = len(frecuencias)
    
            plt.figure(figsize=(9, 5))
            plt.plot(delta_tiempos, frecuencias, marker='o', linestyle='-', label='Frecuencia')
            plt.xlabel('Tiempo relativo [s]', fontsize=12)
            plt.ylabel('Frecuencia [Hz]', fontsize=12)
            plt.title('Frecuencia vs Tiempo relativo')
            plt.grid(True, which='both', linestyle='--', alpha=0.5)
            plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))
    
            texto_stats = (f"Máx: {maximo:.3f} Hz\n"
                        f"Mín: {minimo:.3f} Hz\n"
                        f"Media: {media:.3f} Hz\n"
                        f"Mediana: {mediana:.3f} Hz\n"
                        f"Nº puntos: {n_puntos}")
            plt.gca().text(0.98, 0.02, texto_stats, fontsize=10,
                        ha='right', va='bottom', transform=plt.gca().transAxes,
                        bbox=dict(facecolor='white', alpha=0.7, edgecolor='gray'))
            plt.tight_layout()
            plt.show()
    
        # ====== SECCIÓN 8: Exportar a Excel opcional ======
        if exportar_excel:
            import pandas as pd
            from datetime import datetime
    
            fecha_hora = datetime.now().strftime("%S_%M_%H_%d_%m_%Y")
            raw_data = {
                "Muestra": [f"Muestra{i}" for i in range(len(frecuencias))],
                "Frecuencia [Hz]": np.round(frecuencias, 6),
                "Timestamp [s]": np.round(timestamps_utiles, 6),
                "Delta_tiempo [s]": np.round(delta_tiempos, 6)
            }
            df_raw = pd.DataFrame(raw_data)
            nombre_raw = f"RawDataFreqYTiempo_BTBack_{fecha_hora}.xlsx"
            df_raw.to_excel(nombre_raw, index=False, float_format="%.6f")
            print(f"Archivo de datos crudos guardado como: {nombre_raw}")
    
        # ====== SECCIÓN 9: Devolver resultados ======
        return frecuencias, timestamps_utiles, delta_tiempos
    

    

    def medir_n_muestras_equidistantesBTBack2(
        self,
        n_muestras=100,
        canal='A',
        intervalo_captura=0.2,
        graficarFT=False,
        exportar_excel=False
    ):
        """
        Adquisición rápida de un array de frecuencias usando los comandos:
        - :MEASure:ARRay:FREQuency:BTBack? N,canal
    
        Permite especificar el tiempo de integración (apertura) para cada medición.
    
        Parámetros de entrada:
        ----------------------
        n_muestras : int
            Número de muestras a adquirir (default: 100)
        canal : str o int
            Canal de medida: 'A', 'B', 1 o 2 (default: 'A')
        intervalo_captura : float o None
            Tiempo de integración (apertura) en segundos para cada muestra.
            - Mínimo típico: 4e-6 s (4 µs)
            - Máximo típico: 1000 s (depende del instrumento)
            - Default CNT-91: 0.2 s
            Si se pasa None, NO se configura el tiempo de integración y se usa el valor actual del instrumento.
        graficarFT : bool
            Si True, grafica frecuencia vs tiempo relativo (default: False)
        exportar_excel : bool
            Si True, exporta los datos a un archivo Excel (default: False)
    
        Salida:
        -------
        frecuencias : np.ndarray
            Array de frecuencias medidas (Hz)
        tiempos_relativos : np.ndarray
            Array de tiempos relativos (s) generados por software
        """
    
        import numpy as np
        import time
    
        # ====== Validación de parámetros y canal ======
        canales = {'A': 1, 'B': 2, '1': 1, '2': 2}
        ch = str(canal).upper()
        if ch not in canales:
            raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
        canal_num = canales[ch]
    
        # ====== Reset y limpieza del instrumento ======
        self.dev.write('*RST')
        self.dev.write('*CLS')
     
        # ====== Configuración del tiempo de integración (apertura) ======
        if intervalo_captura is not None:
            self.dev.write(f"SENS:ACQ:APER {intervalo_captura}")
    
        # ====== Adquisición de frecuencias ======
        self.dev.write(f":MEASure:ARRay:FREQuency:BTBack? {n_muestras},{canal_num}")
        time.sleep(n_muestras*intervalo_captura*1.3)
        data_freq = self.dev.read()
        frecuencias = np.array([float(val) for val in data_freq.strip().split(',') if val])
    
        if len(frecuencias) != n_muestras:
            raise RuntimeError("El número de frecuencias adquiridas no coincide con el solicitado.")
    
        # ====== Generación del eje temporal sintético ======
        tiempos_relativos = np.arange(len(frecuencias)) * intervalo_captura
    
        # ====== Visualización opcional ======
        if graficarFT:
            import matplotlib.pyplot as plt
            from matplotlib.ticker import MaxNLocator
    
            maximo = np.max(frecuencias)
            minimo = np.min(frecuencias)
            media = np.mean(frecuencias)
            mediana = np.median(frecuencias)
            n_puntos = len(frecuencias)
    
            plt.figure(figsize=(9, 5))
            plt.plot(tiempos_relativos, frecuencias, marker='o', linestyle='-', label='Frecuencia')
            plt.xlabel('Tiempo relativo [s]', fontsize=12)
            plt.ylabel('Frecuencia [Hz]', fontsize=12)
            plt.title('Frecuencia vs Tiempo relativo')
            plt.grid(True, which='both', linestyle='--', alpha=0.5)
            plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))
    
            texto_stats = (f"Máx: {maximo:.3f} Hz\n"
                           f"Mín: {minimo:.3f} Hz\n"
                           f"Media: {media:.3f} Hz\n"
                           f"Mediana: {mediana:.3f} Hz\n"
                           f"Nº puntos: {n_puntos}")
            plt.gca().text(0.98, 0.02, texto_stats, fontsize=10,
                           ha='right', va='bottom', transform=plt.gca().transAxes,
                           bbox=dict(facecolor='white', alpha=0.7, edgecolor='gray'))
            plt.tight_layout()
            plt.show()
    
        # ====== Exportar a Excel opcional ======
        if exportar_excel:
            import pandas as pd
            from datetime import datetime
    
            fecha_hora = datetime.now().strftime("%S_%M_%H_%d_%m_%Y")
            raw_data = {
                "Muestra": [f"Muestra{i}" for i in range(len(frecuencias))],
                "Frecuencia [Hz]": np.round(frecuencias, 6),
                "Tiempo relativo [s]": np.round(tiempos_relativos, 6)
            }
            df_raw = pd.DataFrame(raw_data)
            nombre_raw = f"RawDataFreqYTiempo_BTBack_{fecha_hora}.xlsx"
            df_raw.to_excel(nombre_raw, index=False, float_format="%.6f")
            print(f"Archivo de datos crudos guardado como: {nombre_raw}")
    
        # ====== Devolver resultados ======
        return frecuencias, tiempos_relativos





    def allan_deviation_vs_tau_JOAN(self, vector_t=None, vector_freq=None, plot=False):
            """
            Calcula la desviación de Allan usando datos guardados en el objeto o los que le pases.
            """
            import numpy as np
            import matplotlib.pyplot as plt
    
            # Si no se pasan argumentos, usa los guardados en el objeto
            if vector_t is None:
                vector_t = self.vector_t
            if vector_freq is None:
                vector_freq = self.vector_freq
    
            N = len(vector_freq)
            f0 = np.mean(vector_freq)
            tsample = np.mean(np.diff(vector_t))
            fsample = 1/tsample
            freq_cumsum = np.cumsum(vector_freq) / fsample
            vector_tau = tsample * np.arange(1, int(N/2) + 1)
            ntau = len(vector_tau)
            sigmay2 = np.zeros(ntau)
            for ii in range(ntau-1):
                vector_index = np.arange(0, N, ii+1)
                M = len(vector_index)
                x = freq_cumsum[vector_index] / f0
                y = (x[1:M] - x[0:M-1]) / vector_tau[ii]
                sigmay2[ii] = np.mean((y[1:M-1] - y[0:M-2])**2) / 2
            if plot:
                fig = plt.figure()
                plt.loglog(vector_tau, np.sqrt(sigmay2), 'r', linewidth=1.5)
                plt.grid(True)
                plt.xlabel('Tiempo de integración (τ)')
                plt.ylabel('Desviación de Allan (σ)')
                plt.title('Desviación de Allan')
            return vector_tau, sigmay2







    def medir_n_muestras_equidistantesBTBack3(
            self,
            Exp_Value,
            n_muestras=100,
            canal='A',
            intervalo_captura=0.2,
            graficarFT=False,
            exportar_excel=False
        ):
        """
        Adquisición de n_muestras de períodos back-to-back y conversión a frecuencias,
        garantizando que el tiempo de integración (apertura) se aplique correctamente.
    
        Parámetros:
        -----------
        n_muestras : int
            Número de muestras de periodo a adquirir.
        canal : 'A'|'B'|1|2
            Canal de medida.
        intervalo_captura : float o None
            Tiempo de integración (s). Si None, no se toca y se usa valor actual.
        graficarFT : bool
            Si True, grafica frecuencia vs tiempo relativo.
        exportar_excel : bool
            Si True, exporta a Excel.
    
        Devuelve:
        ---------
        frecuencias : np.ndarray
            Frecuencias calculadas (Hz).
        tiempos_periodo : np.ndarray
            Tiempos absolutos de cada medición de período (s).
        delta_tiempos : np.ndarray
            Tiempos relativos al inicio (s).
        """
    
        import numpy as np
        import time
        # 1) Canal
        canales = {'A':1, 'B':2, '1':1, '2':2}
        ch = str(canal).upper()
        if ch not in canales:
            raise ValueError("Canal debe ser 'A','B',1 o 2")
        canal_num = canales[ch]
    
        # 2) Reset inicial
        self.dev.write("*RST")
        self.dev.query("*OPC?")
        self.dev.write("*CLS")
        # Asegurar timestamps ON (aunque aquí no usamos TSTAmp)
        self.dev.write(":FORMat:TINFormation ON")
    
        # 3) Apertura & pacing para PERIOD
        if intervalo_captura is not None:
            # Apertura para Period
            self.dev.write(f':SENS:ACQ:APER {intervalo_captura}')
            
            # Pacing time
            self.dev.write(f":TRIG:TIM {intervalo_captura}")
            

    
        # 4) Lectura de períodos BTBack
        # Devuelve un array de N periodos en segundos
        cmd = f":MEASure:ARRay:PERiod:BTBack? {n_muestras},({canal_num})"
        self.dev.write(cmd)
        time.sleep(n_muestras*intervalo_captura*1.4)
        raw = self.dev.read().strip().split(',')
        periodos = np.array([float(v) for v in raw if v], dtype=float)
        
        # 5) Timestamps asociados: utilizamos la misma llamada a TSTAmp si queremos tiempos
        self.dev.write(f":MEASure:ARRay:TSTAmp? {n_muestras},({canal_num})")
        time.sleep(50)
        raw_ts = self.dev.read().strip().split(',')
        ts_all = np.array([float(v) for v in raw_ts if v], dtype=float)
        # TSTAmp devuelve pares, tomamos los "ends" (índices 1,3,5...)
        ts_ends = ts_all[1::2]
        
        if len(ts_ends) != len(periodos):
            raise RuntimeError("Número de timestamps y periodos no coincide.")
    
        # 6) Conversión periodo→frecuencia
        frecuencias = 1.0 / periodos
        delta_t = ts_ends - ts_ends[0]
    
        # 7) Gráfica opcional
        if graficarFT:
            import matplotlib.pyplot as plt
            from matplotlib.ticker import MaxNLocator
    
            plt.figure(figsize=(9,5))
            plt.plot(delta_t, frecuencias, marker='o', linestyle='-')
            plt.xlabel('Tiempo relativo [s]')
            plt.ylabel('Frecuencia [Hz]')
            plt.title('Frecuencia (1/Periodo) vs Tiempo relativo')
            plt.grid(True, which='both', linestyle='--', alpha=0.5)
            plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))
    
            stats = (
                f"Máx: {frecuencias.max():.3f} Hz\n"
                f"Mín: {frecuencias.min():.3f} Hz\n"
                f"Media: {frecuencias.mean():.3f} Hz\n"
                f"Mediana: {np.median(frecuencias):.3f} Hz\n"
                f"N pts: {len(frecuencias)}"
            )
            plt.gca().text(
                0.98, 0.02, stats,
                ha='right', va='bottom',
                transform=plt.gca().transAxes,
                bbox=dict(facecolor='white', alpha=0.7, edgecolor='gray')
            )
            plt.tight_layout()
            plt.show()
    
        # 8) Exportar a Excel opcional
        if exportar_excel:
            import pandas as pd
            from datetime import datetime
    
            now = datetime.now().strftime("%Y%m%d_%H%M%S")
            df = pd.DataFrame({
                "Muestra": [f"P{i}" for i in range(len(periodos))],
                "Periodo [s]": np.round(periodos, 9),
                "Frecuencia [Hz]": np.round(frecuencias, 6),
                "Timestamp [s]": np.round(ts_ends, 6),
                "Delta_t [s]": np.round(delta_t, 6)
            })
            fname = f"DatosPeriodo_BTBack_{now}.xlsx"
            df.to_excel(fname, index=False, float_format="%.6f")
            print(f"Archivo guardado: {fname}")
    
        return frecuencias, ts_ends, delta_t
        

    def allan_deviation_tau0_JAUME(self, vector_t=None, vector_freq=None):
        """
        Calcula la desviación de Allan únicamente para τ = tsample,
        es decir, el intervalo de muestreo inicial.
        Devuelve (tau, adev) donde:
          - tau: tiempo de integración (s)
          - adev: desviación de Allan σ(τ)
        """
        import numpy as np
    
        # Si no se pasan, uso los vectores del objeto
        if vector_t    is None: vector_t    = self.vector_t
        if vector_freq is None: vector_freq = self.vector_freq
    
        N = len(vector_freq)
        if N < 3:
            raise ValueError("Se necesitan al menos 3 muestras para calcular Allan deviation.")
    
        # Frecuencia media y periodo de muestreo
        f0      = np.mean(vector_freq)
        tsample = np.mean(np.diff(vector_t))
    
        # Integración de la frecuencia -> fase
        fsample      = 1.0/tsample
        freq_cumsum  = np.cumsum(vector_freq) / fsample
    
        # Agrupamiento m = 1 → índice espaciado de 1 en 1
        idx = np.arange(0, N, 1)
        x   = freq_cumsum[idx] / f0
    
        # Diferencias normalizadas de fase → frecuencia media en cada intervalo
        # hay N-1 valores de y
        y = (x[1:] - x[:-1]) / tsample
    
        # Allan variance para τ = tsample: varianza de las diferencias de y
        sig2 = np.mean((y[1:] - y[:-1])**2) / 2.0
    
        # Allan deviation σ(τ)
        adev = np.sqrt(sig2)
    
        return tsample, adev



# ******************************** DATALOGGER ******************************** #






    def configurar_dispositivo(self,
                               canal='A',
                               intervalo_s=0.2,
                               acoplamiento='AC',
                               impedancia='MIN',
                               atenuacion='1',
                               trigger_level=None,
                               trigger_slope='POS',
                               filtro_Digital_PASSAbaja=None,
                               filtro_Analog_PASSAbaja=None,
                               file_path=None):
        """
        Configura el dispositivo según parámetros dados y guarda la configuración en un archivo Excel
        dentro de la subcarpeta "Datalogger_Mediciones".
    
        - Si file_path es None, genera un nombre con fecha y hora actuales.
        - Crea carpeta si no existe.
        - Escribe hoja "Configuración" como una tabla Excel,
          mapeando valores según normas:
          * Impedancia: 'MAX'->'1Mohm', 'MIN'->'50ohm'
          * Atenuación: concatena 'x'
          * Trigger Level: si None, '70%' (canal A/1) o '30%' (canal B/2)
          * Filtro Digital PASAbaja: concatena 'Hz'
          * Filtro Analógico PASAbaja: 'VERDADERO'/'FALSO'
        """
        # === RESET Y LIMPIEZA ===
        self.dev.write('*RST')
        self.dev.write('*CLS')
    
        # === CONFIGURACIÓN DEL DISPOSITIVO ===
        canales = {'A': '@1', 'B': '@2', '1': '@1', '2': '@2'}
        ch = str(canal).upper()
        if ch not in canales:
            raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
        # Enviar comandos básicos
        self.dev.write('DISP:ENAB OFF')
        self.dev.write(f'SENS:ACQ:APER {intervalo_s}')
        canal_12 = 1 if ch in ('A', '1') else 2
        if acoplamiento:
            self.dev.write(f':INP{canal_12}:COUP {acoplamiento}')
        # Mapeo de impedancia
        if impedancia:
            self.dev.write(f':INP{canal_12}:IMP {impedancia}')
        if atenuacion:
            self.dev.write(f':INP{canal_12}:ATT {atenuacion}')
            if atenuacion == 'MAX' and trigger_level is not None:
                trigger_level = float(trigger_level) * 10
        if trigger_level is not None:
            self.dev.write(f':INP{canal_12}:LEV:AUTO OFF')
            self.dev.write(f':INP{canal_12}:LEV {trigger_level}')
        if trigger_slope:
            self.dev.write(f':INP{canal_12}:SLOPe {trigger_slope}')
        if filtro_Digital_PASSAbaja is not None:
            self.dev.write(f':INP{canal_12}:FILT:DIG ON')
            self.dev.write(f':INP{canal_12}:FILT:DIG:FREQ {filtro_Digital_PASSAbaja}')
        if filtro_Analog_PASSAbaja is True:
            self.dev.write(f':INP{canal_12}:FILT ON')
        self.dev.write('DISP:ENAB ON')
    
        # === NOMBRE Y RUTA DEL ARCHIVO ===
        if not file_path:
            now = datetime.now()
            file_path = (
                f"Medición_dia{now.day}_mes{now.month}_año{now.year}_"
                f"hora{now.hour}_min{now.minute}_seg{now.second}.xlsx"
            )
        subfolder = os.path.join(os.getcwd(), 'Datalogger_Mediciones')
        os.makedirs(subfolder, exist_ok=True)
        full_path = os.path.join(subfolder, os.path.basename(file_path))
    
        # === PREPARAR VALORES PARA LA TABLA ===
        # Impedancia: mapear
        if impedancia == 'MAX':
            display_imp = '1Mohm'
        elif impedancia == 'MIN':
            display_imp = '50ohm'
        else:
            display_imp = impedancia
        # Atenuación: concatenar 'x'
        display_att = f"{atenuacion}x"
        # Trigger Level: por defecto si None
        if trigger_level is None:
            display_trig = '70%' if canal_12 == 1 else '30%'
        else:
            display_trig = str(trigger_level)
        # Filtro Digital PASAbaja: si existe, añadir 'Hz'
        if filtro_Digital_PASSAbaja is not None:
            display_filt_dig = f"{filtro_Digital_PASSAbaja}Hz"
        else:
            display_filt_dig = None
        # Filtro Analógico PASAbaja: boolean
        if filtro_Analog_PASSAbaja is True:
            display_filt_ana = 'VERDADERO'
        elif filtro_Analog_PASSAbaja is False:
            display_filt_ana = 'FALSO'
        else:
            display_filt_ana = None
    
        config_values = [
            ("Canal", ch),
            ("Intervalo (s)", intervalo_s),
            ("Acoplamiento", acoplamiento),
            ("Impedancia", display_imp),
            ("Atenuación", display_att),
            ("Trigger Level", display_trig),
            ("Trigger Slope", trigger_slope),
            ("Filtro Digital PASAbaja", display_filt_dig),
            ("Filtro Analógico PASAbaja", display_filt_ana)
        ]
    
        # === CREAR/CARGAR EXCEL Y ESCRIBIR TABLA ===
        if os.path.exists(full_path):
            wb = load_workbook(full_path)
            if "Configuración" in wb.sheetnames:
                wb.remove(wb["Configuración"])
        else:
            wb = Workbook()
        ws = wb.create_sheet("Configuración")
    
        # Encabezados
        ws.cell(row=1, column=1, value="Parámetro").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Valor").font = Font(bold=True)
        # Filas
        for idx, (param, val) in enumerate(config_values, start=2):
            ws.cell(row=idx, column=1, value=param)
            ws.cell(row=idx, column=2, value=val)
        # Tabla
        max_row = len(config_values) + 1
        table = Table(displayName="TablaConfiguracion", ref=f"A1:B{max_row}")
        style = TableStyleInfo(name="TableStyleMedium9",
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
    
        try:
            wb.save(full_path)
        except Exception as e:
            print(f"Error al guardar archivo de configuración: {e}")
            raise
        finally:
            wb.close()  # Cerrar explícitamente el archivo
        
        # dentro de configurar_dispositivo, justo antes de return:
        self.file_path = full_path
        return full_path
            
    
    
    
    
    def start_continuous_measurement(self, intervalo_s=0.2,n_muestras=1, canal='A'):
        """
        Inicia la medición continua en el canal y intervalo indicados.
        Sólo configura el instrumento y lanza INIT, sin leer ni abortar.
        """
        # ========== Validación y selección de canal ==========
        canales = {'A': '@1', 'B': '@2', '1': '@1', '2': '@2'}
        ch = str(canal).upper()
        if ch not in canales:
            raise ValueError("El canal debe ser 'A', 'B', 1 o 2")
        canal_cmd = canales[ch]
    
        # ========== Configuración mínima del instrumento ==========

        self.dev.write('CAL:INT:AUTO OFF')        # Calibración automática off
        self.dev.write('DISP:ENAB OFF')           # Display off
        self.dev.write(f'CONF:FREQ {canal_cmd}')  # Configurar frecuencia en canal
        self.dev.write(f'SENS:ACQ:APER {intervalo_s}')  # Apertura por muestra
        self.dev.write('ARM:COUNT INF')           # Recuento infinito
        self.dev.write('ARM:CONT ON')             # Modo continuo
        self.dev.write('FORM:TINF ON')            # Incluir timestamps
        
        # ========== Lanzamiento de medición continua ==========
        self.dev.write('INIT')
        
        tiempo_espera = 0.25
        
        return tiempo_espera

    def fetch_continuous_samples(self, n_muestras=1,tiempo_espera=1):
        """
        Recupera exactamente n_muestras de la medición continua ya iniciada.
        Devuelve dos listas: frecuencias y timestamps.
        """
        time.sleep(tiempo_espera*n_muestras)
        # Solicita las muestras
        self.dev.write(f'FETC:ARR? {n_muestras}')
        data = self.dev.read()
    
        # Parseo de la respuesta CSV en floats
        valores = [float(val) for val in data.strip().split(',') if val]
    
        # Separar frecuencias (posiciones pares) y timestamps (impares)
        frecuencias = valores[::2]
        timestamps  = valores[1::2]
    
        # Comprobar si llegaron suficientes muestras
        if len(frecuencias) < n_muestras:
            print(f"¡Advertencia! Sólo {len(frecuencias)} muestras recibidas de las {n_muestras} solicitadas.")
    
        return frecuencias, timestamps
    


    def abort_continuous_measurement(self):
        """
        Aborta la medición continua en curso.
        """
        self.dev.write('*RST')          # NO HACE FALTA PERO ES INTERESANTE
        self.dev.write("*CLS")          # NO HACE FALTA PERO ES INTERESANTE
        self.dev.write('ABOR')          # TERMINA LA MEDICIÓN CONTINUA 
        self.dev.write('DISP:ENAB ON')           # Display ON   

    def append_measurement(self, frecuencia, timestamp, hoja="Mediciones"):
        if not hasattr(self, "file_path"):
            raise RuntimeError("Primero configura el dispositivo para generar el archivo Excel.")
        
        try:
            wb = load_workbook(self.file_path)
            if hoja not in wb.sheetnames:
                ws = wb.create_sheet(hoja)
                ws.append(["Timestamp (s)", "Frecuencia (Hz)"])  # encabezados
            else:
                ws = wb[hoja]
            ws.append([timestamp, frecuencia])
            wb.save(self.file_path)
            wb.close()  # Cerrar explícitamente el archivo
        except Exception as e:
            print(f"Error al escribir en Excel: {e}")
            # Intentar cerrar el archivo si está abierto
            try:
                if 'wb' in locals():
                    wb.close()
            except:
                pass
            raise

    def cerrar_archivo_excel(self):
        """
        Cierra de manera segura el archivo Excel si está abierto.
        Este método debe llamarse antes de terminar el programa para evitar
        problemas de archivos bloqueados.
        """
        try:
            if hasattr(self, 'file_path') and self.file_path:
                # Intentar cerrar cualquier archivo Excel que pueda estar abierto
                # Esto es principalmente para limpiar recursos de openpyxl
                import gc
                gc.collect()  # Forzar la recolección de basura
                print(f"Archivo Excel cerrado correctamente: {os.path.basename(self.file_path)}")
        except Exception as e:
            print(f"Advertencia al cerrar archivo Excel: {e}")

    def cerrar_conexion(self):
        """
        Cierra la conexión con el dispositivo CNT-91 de manera segura.
        
        Esta función:
        1. Cierra el archivo Excel si está abierto
        2. Aborta cualquier medición en curso
        3. Resetea el instrumento a su estado por defecto
        4. Limpia el buffer de errores
        5. Cierra la conexión VISA
        6. Libera los recursos del ResourceManager
        
        Es importante llamar a esta función al finalizar el uso del instrumento
        para evitar problemas de comunicación en futuras sesiones.
        
        Ejemplo de uso:
        >>> cnt = CNT_frequenciometro()
        >>> # ... realizar mediciones ...
        >>> cnt.cerrar_conexion()
        """
        try:
            # 1. Cerrar el archivo Excel primero
            self.cerrar_archivo_excel()
            
            # 2. Abortar cualquier medición en curso
            try:
                self.dev.write('ABOR')
            except:
                pass  # Si ya está abortado, no hay problema
            
            # 3. Resetear el instrumento a estado por defecto
            try:
                self.dev.write('*RST')
            except:
                pass
            
            # 4. Limpiar buffer de errores
            try:
                self.dev.write('*CLS')
            except:
                pass
            
            # 5. Reactivar display si estaba desactivado
            try:
                self.dev.write('DISP:ENAB ON')
            except:
                pass
            
            # 6. Cerrar la conexión VISA
            try:
                if hasattr(self, 'dev') and self.dev is not None:
                    self.dev.close()
                    self.dev = None
                    print("Conexión con el dispositivo CNT-91 cerrada correctamente.")
            except Exception as e:
                print(f"Advertencia al cerrar conexión VISA: {e}")
            
            # 7. Liberar recursos del ResourceManager (opcional)
            # Nota: En algunas implementaciones, el ResourceManager se libera automáticamente
            # al finalizar el programa, pero es buena práctica hacerlo explícitamente
            
        except Exception as e:
            print(f"Error durante el cierre de conexión: {e}")
            # Aún intentamos cerrar la conexión VISA si es posible
            try:
                if hasattr(self, 'dev') and self.dev is not None:
                    self.dev.close()
                    self.dev = None
            except:
                pass
        finally:
            # Marcar que la conexión está cerrada
            self._conexion_cerrada = True

    def __del__(self):
        """
        Destructor de la clase. Se ejecuta automáticamente cuando el objeto
        se elimina de la memoria. Asegura que la conexión se cierre correctamente.
        """
        try:
            if hasattr(self, 'dev') and self.dev is not None:
                self.cerrar_conexion()
        except:
            pass  # En el destructor, no queremos que se lancen excepciones

    def verificar_archivo_disponible(self):
        """
        Verifica si el archivo Excel está disponible para escritura.
        Útil para detectar problemas de archivos bloqueados.
        """
        if not hasattr(self, 'file_path') or not self.file_path:
            return False
        
        try:
            # Intentar abrir el archivo en modo escritura para verificar si está bloqueado
            with open(self.file_path, 'a') as f:
                pass
            return True
        except PermissionError:
            print(f"Advertencia: El archivo {os.path.basename(self.file_path)} está siendo usado por otro proceso.")
            return False
        except Exception as e:
            print(f"Error al verificar archivo: {e}")
            return False









#              measure_frequency

"""
# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

"""             
"""

try:
    resultado = objt_prueba.measure_frequency('A')
    print("Resultado de measure_frequency", resultado)
except Exception as e:
    print("Error al ejecutar measure_frequency", e)

"""





#             Measure_temperature_example

"""
# MIDE LA TEMPERATURA

    
    
"""

"""
try:
    temperatura = objt_prueba.Measure_temperature_example()
    print("Resultado de Measure_temperature_example", temperatura)
except Exception as t:
    print("Error al ejecutar Measure_temperature_example", t)

"""



#               measure_frequency_array_CONTINUOUS    

"""

# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)
 # ====== NUEVA SECCIÓN: Prueba de medida continua con try/except ======
try:
    if not resources:
        raise RuntimeError("No hay recursos VISA disponibles.")
    
    duration = 5.0  # Duración de la prueba en segundos
    print(f"\nIniciando medida continua durante {duration:.1f} s en canal A...")
    freqs = objt_prueba.measure_frequency_array_CONTINUOUS(duration_s=duration, channel='A')

    print(f"\nRecibidos {len(freqs)} valores:")
    for idx, f in enumerate(freqs, start=1):
        print(f"  #{idx}: {f:.6f} Hz")
    if not freqs:
        print("  ¡No se recibieron valores! Verifica la conexión y configuración.")

except Exception as e:
    print(f"Error durante la prueba de medida continua: {e}")
    
    
"""


#             medir_n_muestras_equidistantes     DELTA TIEMPOS y Tiempo de espera

"""

# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

# Parámetros de la prueba
n_muestras = 500
intervalo_s = 0.2

# Calcular y mostrar el tiempo de espera antes de medir
tiempo_espera = n_muestras * intervalo_s * 1.1

# Conversión a formato horas:minutos:segundos
horas = int(tiempo_espera // 3600)
minutos = int((tiempo_espera % 3600) // 60)
segundos = tiempo_espera % 60

print(f"TIEMPO DE ESPERA ESTIMADO = {tiempo_espera:.2f} segundos "
      f"({horas:02d}:{minutos:02d}:{segundos:05.2f} [hh:mm:ss])")

# Ejecutar la medición
resultados = objt_prueba.medir_n_muestras_equidistantes(n_muestras=n_muestras, intervalo_s=intervalo_s)

# Ajustar los timestamps para restar el inicial
if resultados and isinstance(resultados[0], tuple) and len(resultados[0]) == 2:
    t0 = resultados[0][1]
    resultados_rel = [(f, t - t0) for (f, t) in resultados]
else:
    resultados_rel = resultados  # Si el formato es inesperado, dejar igual

print("Frecuencia (Hz), Delta t (s) respecto a la primera muestra:")
for freq, dt in resultados_rel:
    print(f"{freq:.6f}, {dt:.6f}")


"""




#             medir_n_muestras_equidistantes     DELTA TIEMPOS

"""
# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

resultados = objt_prueba.medir_n_muestras_equidistantes(n_muestras=20, intervalo_s=0.2)

# Ajustar los timestamps para restar el inicial
if resultados and isinstance(resultados[0], tuple) and len(resultados[0]) == 2:
    t0 = resultados[0][1]
    resultados_rel = [(f, t - t0) for (f, t) in resultados]
else:
    resultados_rel = resultados  # Si el formato es inesperado, dejar igual

print("Frecuencia (Hz), Delta t (s) respecto a la primera muestra:")
for freq, dt in resultados_rel:
    print(f"{freq:.6f}, {dt:.6f}")

"""



#             medir_n_muestras_equidistantes     TIEMPOS


"""
# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de ======



resultados = objt_prueba.medir_n_muestras_equidistantes(n_muestras=20, intervalo_s=0.2)
print(resultados)

"""




#             medir_n_muestras_equidistantes     TIEMPOS y tiempos relativos

"""

# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

# Parámetros de la prueba
n_muestras = 100

intervalo_s = 0.2

# Calcular y mostrar el tiempo de espera antes de medir
tiempo_espera = n_muestras * intervalo_s * 1.1

# Conversión a formato horas:minutos:segundos
horas = int(tiempo_espera // 3600)
minutos = int((tiempo_espera % 3600) // 60)
segundos = tiempo_espera % 60

print(f"TIEMPO DE ESPERA ESTIMADO = {tiempo_espera:.2f} segundos "
      f"({horas:02d}:{minutos:02d}:{segundos:05.2f} [hh:mm:ss])")

# Ejecutar la medición con la nueva función V2
frecuencias, timestamps, delta_tiempos = objt_prueba.medir_n_muestras_equidistantesV2(n_muestras=n_muestras, intervalo_s=intervalo_s)

# Mostrar los resultados en el formato solicitado
print("\nResultados de la medición:")
for i in range(len(frecuencias)):
    print(f"Muestra {i+1}: {frecuencias[i]:.6f} Hz, {timestamps[i]:.6f} s, {delta_tiempos[i]:.6f} s") 

"""








#      medir_n_muestras_equidistantes     TIEMPOS y tiempos relativos y graficar F vs T

"""
# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

# Parámetros de la prueba
n_muestras = 10

intervalo_s = 0.2

# Calcular y mostrar el tiempo de espera antes de medir
tiempo_espera = n_muestras * intervalo_s * 1.1

# Conversión a formato horas:minutos:segundos
horas = int(tiempo_espera // 3600)
minutos = int((tiempo_espera % 3600) // 60)
segundos = tiempo_espera % 60

print(f"TIEMPO DE ESPERA ESTIMADO = {tiempo_espera:.2f} segundos "
      f"({horas:02d}:{minutos:02d}:{segundos:05.2f} [hh:mm:ss])")

# Ejecutar la medición con la nueva función V2
frecuencias, timestamps, delta_tiempos = objt_prueba.medir_n_muestras_equidistantesV3(n_muestras=n_muestras, intervalo_s=intervalo_s)

# Mostrar los resultados en el formato solicitado
print("\nResultados de la medición:")
for i in range(len(frecuencias)):
    print(f"Muestra {i+1}: {frecuencias[i]:.6f} Hz, {timestamps[i]:.6f} s, {delta_tiempos[i]:.6f} s") 


"""




#      medir_n_muestras_equidistantes     TIEMPOS y tiempos relativos y graficar F vs T
# Medir ADEVS



"""
# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

# Parámetros de la prueba
n_muestras = 10
intervalo_s = 0.2

# Calcular y mostrar el tiempo de espera antes de medir
tiempo_espera = n_muestras * intervalo_s * 1.1

# Conversión a formato horas:minutos:segundos
horas = int(tiempo_espera // 3600)
minutos = int((tiempo_espera % 3600) // 60)
segundos = tiempo_espera % 60

print(f"TIEMPO DE ESPERA ESTIMADO = {tiempo_espera:.2f} segundos "
      f"({horas:02d}:{minutos:02d}:{segundos:05.2f} [hh:mm:ss])")

# Ejecutar la medición con la nueva función V4
frecuencias, timestamps, delta_tiempos, allan_deviations, taus = objt_prueba.medir_n_muestras_equidistantesV4(
    n_muestras=n_muestras, intervalo_s=intervalo_s, graficarFT=True,
)

# Mostrar los resultados en el formato solicitado (con unidades)
print("\nResultados de la medición:")
for i in range(len(frecuencias)):
    print(f"Muestra {i+1} : {frecuencias[i]:.6f} Hz, {timestamps[i]:.6f} s, {delta_tiempos[i]:.6f} s")

print("\nDATO  : Allan deviations y Taus")
# Mostrar los pares Allan deviation y Tau juntos y con unidades
for i in range(len(allan_deviations)):
    print(f"Tau {taus[i]:.3f} s: Allan deviation = {allan_deviations[i]:.6f} Hz")

"""



"""
# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

# Parámetros de la prueba
n_muestras = 500
intervalo_s = 0.5

# Calcular y mostrar el tiempo de espera antes de medir
tiempo_espera = n_muestras * intervalo_s * 1.1

# Conversión a formato horas:minutos:segundos
horas = int(tiempo_espera // 3600)
minutos = int((tiempo_espera % 3600) // 60)
segundos = tiempo_espera % 60

print(f"TIEMPO DE ESPERA ESTIMADO = {tiempo_espera:.2f} segundos "
      f"({horas:02d}:{minutos:02d}:{segundos:05.2f} [hh:mm:ss])")

# Ejecutar la medición con la nueva función V4
frecuencias, timestamps, delta_tiempos, allan_deviations, taus = objt_prueba.medir_n_muestras_equidistantesV6(
    n_muestras=n_muestras, intervalo_s=intervalo_s, graficarFT=True,graficarDevTau=True,
)

# Mostrar los resultados en el formato solicitado (con unidades)
print("\nResultados de la medición:")
for i in range(len(frecuencias)):
    print(f"Muestra {i+1} : {frecuencias[i]:.6f} Hz, {timestamps[i]:.6f} s, {delta_tiempos[i]:.6f} s")

print("\nDATOS  : Allan deviations y Taus")
# Mostrar los pares Allan deviation y Tau juntos y con unidades
for i in range(len(allan_deviations)):
    print(f"Tau {taus[i]:.3f} s: Allan deviation = {allan_deviations[i]:.6f} Hz")



"""





#      medir_n_muestras_equidistantes     TIEMPOS y tiempos relativos y graficar F vs T
# Medir ADEVS
# guardar valores en excel



"""

# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

# Parámetros de la prueba
n_muestras = 1000
intervalo_s = 0.5

# Calcular y mostrar el tiempo de espera antes de medir
tiempo_espera = n_muestras * intervalo_s * 1.1

# Conversión a formato horas:minutos:segundos
horas = int(tiempo_espera // 3600)
minutos = int((tiempo_espera % 3600) // 60)
segundos = tiempo_espera % 60

print(f"TIEMPO DE ESPERA ESTIMADO = {tiempo_espera:.2f} segundos "
      f"({horas:02d}:{minutos:02d}:{segundos:05.2f} [hh:mm:ss])")

# Ejecutar la medición con la nueva función V4
frecuencias, timestamps, delta_tiempos, allan_deviations, taus = objt_prueba.medir_n_muestras_equidistantesV5(
    n_muestras=n_muestras, intervalo_s=intervalo_s, graficarFT=True,graficarDevTau=True,
)

# Mostrar los resultados en el formato solicitado (con unidades)
print("\nResultados de la medición:")
for i in range(len(frecuencias)):
    print(f"Muestra {i+1} : {frecuencias[i]:.6f} Hz, {timestamps[i]:.6f} s, {delta_tiempos[i]:.6f} s")

print("\nDATO  : Allan deviations y Taus")
# Mostrar los pares Allan deviation y Tau juntos y con unidades
for i in range(len(allan_deviations)):
    print(f"Tau {taus[i]:.3f} s: Allan deviation = {allan_deviations[i]:.6f} Hz")




"""


# Configuración dispositivo

"""
    
    
    cnt = CNT_frequenciometro()
    cnt.configurar_medicion_estadistica(
        tiempo_apertura=10,
        acoplamiento='DC',
        impedancia='50',
        atenuador='10',
        filtro_analogico=True,
        filtro_digital=True,
        freq_filtro_digital=1000,
        nivel_auto=False,
        nivel_disparo=0.3,
        medicion_continua=True,
        numero_muestras=500
    )
    
    
    
    
"""

#

"""

# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

# Parámetros de la prueba
n_muestras = 200

intervalo_s = 2E-4

# Calcular y mostrar el tiempo de espera antes de medir
tiempo_espera = n_muestras * intervalo_s * 1.1

# Conversión a formato horas:minutos:segundos
horas = int(tiempo_espera // 3600)
minutos = int((tiempo_espera % 3600) // 60)
segundos = tiempo_espera % 60

print(f"TIEMPO DE ESPERA ESTIMADO = {tiempo_espera:.2f} segundos "
      f"({horas:02d}:{minutos:02d}:{segundos:05.2f} [hh:mm:ss])")

# Ejecutar la medición con la nueva función V2
frecuencias, timestamps, delta_tiempos = objt_prueba.medir_n_muestras_equidistantesV31(n_muestras=n_muestras, intervalo_s=intervalo_s)

# Mostrar los resultados en el formato solicitado
print("\nResultados de la medición:")
for i in range(len(frecuencias)):
    print(f"Muestra {i+1}: {frecuencias[i]:.6f} Hz, {timestamps[i]:.6f} s, {delta_tiempos[i]:.6f} s") 




"""









#             medir_n_muestras_equidistantes     ADEV interno

"""
# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

adev = objt_prueba.leer_adev_cnt91()
print("Allan deviation interna CNT-91:", adev)
    
   
"""







# Hardware pacing


"""
# ==== PRUEBA DE LA FUNCIÓN medir_n_muestras_equidistantes_hardware ====

# 1. Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# 2. Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# 3. Parámetros de la prueba
n_muestras = 200
intervalo_s = 0.2

# 4. Calcular y mostrar el tiempo de espera antes de medir (ajusta si lo ves necesario)
tiempo_espera = n_muestras * intervalo_s * 1.05  # Alineado con la función hardware
horas = int(tiempo_espera // 3600)
minutos = int((tiempo_espera % 3600) // 60)
segundos = tiempo_espera % 60

print(f"TIEMPO DE ESPERA ESTIMADO = {tiempo_espera:.2f} segundos "
      f"({horas:02d}:{minutos:02d}:{segundos:05.2f} [hh:mm:ss])")

# 5. Ejecutar la medición con la función hardware pacing
frecuencias, timestamps, delta_tiempos = objt_prueba.medir_n_muestras_equidistantes_hardware(
    n_muestras=n_muestras,
    intervalo_s=intervalo_s,
    graficarFT=True,
    exportar_excel=True
)

# 6. Mostrar los resultados en el formato solicitado
print("\nResultados de la medición:")
for i in range(len(frecuencias)):
    print(f"Muestra {i+1}: {frecuencias[i]:.6f} Hz, {timestamps[i]:.6f} s, {delta_tiempos[i]:.6f} s")
"""



# LEER EL ADEV ESTADÍSTICO

"""
# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

adev = objt_prueba.leer_adev_cnt91()
print("Allan deviation interna CNT-91:", adev)
    
   
"""

#PARA FINALIZAR LA CONEXIÓN CON EL DISPOSITIVO

"""
FINALIZAR LA COMUNICACIÓN CON EL DISPOSITIVO (FALTA)

"""

"""
try:
     objt_prueba.Reset_Device()
    
except Exception as t:
    print("Error al ejecutar Measure_temperature_example", t)
    
    
"""   


# PARA LOS ERRORES

"""

try:
    NumeroError = objt_prueba.System_Error_Number()
    print("Resultado de ERROR:", NumeroError)
except Exception as p:
    print("Error al ejecutar Measure_example:", p) 
"""   




"""

# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

# Parámetros de la prueba
n_muestras = 100
intervalo_s = 4e-6

# Calcular y mostrar el tiempo de espera antes de medir
tiempo_espera = n_muestras * intervalo_s * 1.1

# Conversión a formato horas:minutos:segundos
horas = int(tiempo_espera // 3600)
minutos = int((tiempo_espera % 3600) // 60)
segundos = tiempo_espera % 60

print(f"TIEMPO DE ESPERA ESTIMADO = {tiempo_espera:.2f} segundos "
      f"({horas:02d}:{minutos:02d}:{segundos:05.2f} [hh:mm:ss])")

# Ejecutar la medición con la nueva función V4
frecuencias, timestamps, delta_tiempos, allan_deviations, taus = objt_prueba.medir_n_muestras_equidistantesV7(
    n_muestras=n_muestras, intervalo_s=intervalo_s, graficarFT=True,graficarDevTau=True,
)

# Mostrar los resultados en el formato solicitado (con unidades)
print("\nResultados de la medición:")
for i in range(len(frecuencias)):
    print(f"Muestra {i+1} : {frecuencias[i]:.6f} Hz, {timestamps[i]:.6f} s, {delta_tiempos[i]:.6f} s")

print("\nDATOS  : Allan deviations y Taus")
# Mostrar los pares Allan deviation y Tau juntos y con unidades
for i in range(len(allan_deviations)):
    print(f"Tau {taus[i]:.6f} s: Allan deviation = {allan_deviations[i]:.6f} Hz")




"""



"""



# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

# Parámetros de la prueba
n_muestras = 10



intervalo_s = 4.00E-06



# Calcular y mostrar el tiempo de espera antes de medir
tiempo_espera = n_muestras * intervalo_s * 1.1

# Conversión a formato horas:minutos:segundos
horas = int(tiempo_espera // 3600)
minutos = int((tiempo_espera % 3600) // 60)
segundos = tiempo_espera % 60

print(f"TIEMPO DE ESPERA ESTIMADO = {tiempo_espera:.2f} segundos "
      f"({horas:02d}:{minutos:02d}:{segundos:05.2f} [hh:mm:ss])")

# Ejecutar la medición con la nueva función V2
frecuencias, timestamps, delta_tiempos = objt_prueba.continuous_measurament_v31(n_muestras=n_muestras, intervalo_s=intervalo_s)

# Mostrar los resultados en el formato solicitado
print("\nResultados de la medición:")
for i in range(len(frecuencias)):
    print(f"Muestra {i+1}: {frecuencias[i]:.6f} Hz, {timestamps[i]:.6f} s, {delta_tiempos[i]:.6f} s") 





"""


"""

# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

try:
        adev = objt_prueba.leer_adev_cnt91()
        if adev is not None:
            print(f"ADEV obtenida: {adev}")
        else:
            print("Error: no se pudo obtener la ADEV.")
            
except Exception as e:
        print(f"Excepción al probar leer_adev_cnt91: {e}")




"""





"""

# Crear un objeto de la Libreria CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# ====== NUEVA SECCIÓN: Prueba de la función modificada ======

try:
        adev, valor_medio, desviacion_tipica, valor_minimo, valor_maximo = objt_prueba.calcular_Adev_Estadistics(canal='A')
 
            
except Exception as e:
        print(f"Excepción al probar leer_adev_cnt91: {e}")



# Imprime los resultados, incluyendo unidades
print(f"Allan deviation:       {adev} Hz")
print(f"Mean value:            {valor_medio} Hz")
print(f"Standard deviation:    {desviacion_tipica} Hz")
print(f"Minimum value:         {valor_minimo} Hz")
print(f"Maximum value:         {valor_maximo} Hz")


"""




"""


# ====== NUEVA SECCIÓN: Prueba de la función modificada, con configuración SCPI óptima ======

# Crear un objeto de la Librería CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# (1) Abrir conexión VISA al dispositivo (usar el primer recurso listado)
direccion_dispositivo = resources[0]
objt_prueba.dev = rm.open_resource(direccion_dispositivo)

# Función auxiliar para formatear Tau con 1 ó 2 cifras significativas
def _format_tau(tau: float) -> str:
    
    #- Si tau < 1e-3 → 1 cifra significativa
    #- Si tau >= 1e-3 → 2 cifras significativas
    
    if tau < 1e-3:
        return f"{tau:.2g}"
    else:
        return f"{tau:.3g}"

# ========== SECCIÓN OPCIONAL: Parámetros SCPI "óptimos" para estabilidad ==========
# Basado en las pruebas previas, escogemos:
#   - acoplamiento = 'DC'      → Para asegurar que la señal no se rechace si es contínua
#   - impedancia   = '50'      → Entrada a 50 Ω para adaptarse a fuentes estándar
#   - atenuacion   = 1         → Atenuación mínima (1×) para no degradar amplitud
#   - trigger_level= 0.5       → Nivel de trigger en 0.5 V (típico para TTL/CMOS)
#   - trigger_slope= 'POS'     → Disparo en flanco positivo
#   - filtro       = 100e3     → Filtro digital a 100 kHz para rechazar ruido alto
#
# Estos valores aseguran que el contador opere en DC, con adaptación 50 Ω, 
# sin atenuación extra, y un filtro que no sea demasiado ancho ni demasiado estrecho.



 
# (2) SACAMOS LA FREQUENCIA DE REFERENCIA

try:
    resultado = objt_prueba.measure_frequency('A')
    print("Resultado de measure_frequency", resultado)
except Exception as e:
    print("Error al ejecutar measure_frequency", e)

# (3) Llamar a la función mejorada, incluyendo SCPI de entrada
resultados = objt_prueba.calcular_Adev_Estadistics_improved(
        canal='A',
        N_muestras=2, # Máximo 100 mínimo 2 ya que sino no sacas el ADEV
        intervalo_captura_min=2E-6,
        intervalo_captura_max=1, ## esta en segundos
        pasos=1, #[ De momento la separácion parece estar bien pero se podría hacer cada decada]
        pacing_time = 2E-6, #esta en s between 2 micro and 500 s [ Permite medir más rápidamente parece no afectar a la medida] 
        acoplamiento   = 'AC',  # 'DC' or 'AC'
        impedancia = 'MIN',     # 'MAX' or 'MIN'
        atenuacion = 1,         # 1 or 10
        trigger_level=0.2,      # El trigger = trigger_level*atenuacion ; solo valores entre 5 a -5 con 2.5mV de pasos y si atenuacion a 10 será como si fuese 50 a -50 a pasos de 25mV
        trigger_slope= 'NEG',    # 'POS' o 'NEG'
        filtro_Digital_PASSAbaja = '10',      # Default value: 100e3 Hz   , Posible valores: between [1 to 50e6 Hz] or MAX or MIN 
        filtro_Analog_PASSAbaja  = 'ON'  # Cualquier cosa q ponga distinta de None activa el passa bajas 100Khz analogico
        
)

# (4) Mostrar resultados por cada tau (tiempo de integración), usando formateo
for tau, adev, media, desv_tip, minimo, maximo in resultados:
    tau_str = _format_tau(tau)
    adevrelation = adev/(float(resultado))
    print(f"Tau = {tau_str} s:")
    print(f"    Allan Deviation = {adev}")
    print(f"    Estabilidad ADEV = {adevrelation}")
    print(f"    Media           = {media}")
    print(f"    Desv. Típica    = {desv_tip}")
    print(f"    Valor Mínimo    = {minimo}")
    print(f"    Valor Máximo    = {maximo}")
    print()


"""


"""

# ========= CÓDIGO DE PRUEBA Y VISUALIZACIÓN =========

# Crear un objeto de la Librería CNT_9X_pendulum
import CNT_9X_pendulum as CNT
objt_prueba = CNT.CNT_frequenciometro()

# Ver la lista de dispositivos en el GPIB
import pyvisa
rm = pyvisa.ResourceManager()
resources = rm.list_resources()
print("Available VISA resources:", resources)

# (1) Abrir conexión VISA al dispositivo (usar el primer recurso listado)
direccion_dispositivo = resources[0]
objt_prueba.dev = rm.open_resource(direccion_dispositivo)

# Función auxiliar para formatear Tau con 1 ó 2 cifras significativas
def _format_tau(tau: float) -> str:
    
   # - Si tau < 1e-3 → 1 cifra significativa
   # - Si tau >= 1e-3 → 2 cifras significativas
    
    if tau < 1e-3:
        return f"{tau:.2g}"
    else:
        return f"{tau:.3g}"




# (2) Llamar a la función mejorada, incluyendo SCPI de entrada
resultados = objt_prueba.calc_Adev_Estadistics1(
            canal='A',
            N_muestras=100,
            intervalo_captura_min=1E-5,
            intervalo_captura_max=100,
            pasos=70,
            pacing_time=2E-6,
            acoplamiento='AC',
            impedancia='MIN',
            atenuacion='1',
            trigger_level=None,
            trigger_slope='POS',
            filtro_Digital_PASSAbaja=None,
            filtro_Analog_PASSAbaja=None,
            guardar=True          
)
print("Resultados de la función mejorada (diccionario por tau):")
# (3) Mostrar resultados por cada tau (tiempo de integración), usando formateo
for res in resultados:
    tau_str = _format_tau(res["intervalo_captura"])
    print(f"Tau = {tau_str} s:")
    print(f"    Allan Deviation = {res['allan_deviation']}")
    print(f"    Estabilidad ADEV = {res['Estabilidad_Adev']}")
    print(f"    frequencia de la medicion = {res['frequencia_of_medicion']}")
    print(f"    Media           = {res['valor_medio']}")
    print(f"    Desv. Típica    = {res['desviacion_tipica']}")
    print(f"    Valor Mínimo    = {res['valor_minimo']}")
    print(f"    Valor Máximo    = {res['valor_maximo']}")

    print()

"""





